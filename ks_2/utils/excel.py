import convertapi
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.worksheet.table import Table, TableStyleInfo
from aspose.cells import Workbook, HtmlSaveOptions, SaveFormat, License
from bs4 import BeautifulSoup
import aspose.cells as cells
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime
from PyPDF2 import PdfReader, PdfWriter
import os


def html_to_excel():
    html_file = "ks_2/utils/updated_file.html"

    html_load_options = cells.HtmlLoadOptions()

    workbook = cells.Workbook(html_file, html_load_options)

    workbook.save("ks_2/utils/output.xlsx")


def change_html(count_rows):
    with open("ks_2/utils/Счет_на_оплату.html", "r",
              encoding="utf-8") as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, "html.parser")

    element_to_remove = soup.find("div",
                                  string=lambda text: text and "Evaluation Only. Created with Aspose.Cells" in text)

    if element_to_remove:
        element_to_remove.decompose()

    target_row = soup.find("td", string=lambda
        text: text and "Bonolit" in text)

    if target_row:
        parent_row = target_row.find_parent("tr")

        if parent_row:
            num_rows_to_add = count_rows - 1

            if num_rows_to_add >= 1:
                for _ in range(num_rows_to_add):
                    new_row = BeautifulSoup(str(parent_row), "html.parser")

                    parent_row.insert_after(new_row)

            with open("ks_2/utils/updated_file.html",
                      "w",
                      encoding="utf-8") as file:
                file.write(str(soup))


def excel_to_html():
    file_path = "ks_2/utils/КС-2 №1 от 03.02.2025 г..xlsx"
    workbook = Workbook(file_path)

    save_options = HtmlSaveOptions()
    save_options.export_active_worksheet_only = True
    save_options.export_images_as_base64 = True

    html_file = "ks_2/utils/Счет_на_оплату.html"
    workbook.save(html_file, save_options)


def create_ks2_excel(data, formset_data, pdf=False, watch_document=False):
    excel_to_html()
    change_html(len(formset_data))
    html_to_excel()

    file_path = 'ks_2/utils/output.xlsx'

    workbook = openpyxl.load_workbook(file_path)
    workbook.remove(workbook["Evaluation Warning"])

    sheet = workbook["КС-2"]

    for row in sheet.iter_rows():
        if row[0].row in [9, 24, 17]:
            sheet.row_dimensions[row[0].row].height = 22

    start_col = 208
    num_cols = 790

    sheet.delete_cols(start_col, num_cols)

    if data["investor"]:
        sheet['K5'] = f'{data["investor"].naming}, {data["investor"].address}'
    else:
        sheet['K5'] = ''

    sheet['Y7'] = f'{data["counterparty"].naming}, {data["counterparty"].address}'
    sheet['Y9'] = f'{data["organization"].naming}, {data["organization"].address}'
    sheet['EW8'] = ''

    if data["name_construction"]:
        sheet['K11'] = f'{data["name_construction"]}'
    elif data["address_construction"]:
        sheet['K11'] = f'{data["address_construction"]}'
    elif data["name_construction"] and data["address_construction"]:
        sheet['K11'] = f'{data["name_construction"]}, {data["address_construction"]}'
    else:
        sheet['K11'] = ''

    if data["name_object"]:
        sheet['K13'] = f'{data["name_object"]}'
    else:
        sheet['K13'] = ''

    if data["view_okdp"]:
        sheet['EW14'] = f'{data["view_okdp"]}'
    else:
        sheet['EW14'] = ''

    if data["number_agreement"]:
        sheet['EW16'] = f'{data["number_agreement"]}'
    else:
        sheet['EW16'] = ''

    if data['date_agreement']:
        date_str = str(data['date_agreement'])
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        formatted_date = date_obj.strftime("%d %m %Y").split(' ')
        sheet['EW17'] = f'{formatted_date[0]}'
        sheet['FA17'] = f'{formatted_date[1]}'
        sheet['FF17'] = f'{formatted_date[2]}'
    else:
        sheet['EW17'] = ''
        sheet['FA17'] = ''
        sheet['FF17'] = ''

    sheet['BJ20'] = f'{data["name"]}'

    date_str = str(data['date'])
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d-%m-%Y")
    sheet['CB20'] = f'{formatted_date}'

    if data['period_from']:
        date_str = str(data['period_from'])
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        formatted_date = date_obj.strftime("%d-%m-%Y")

        sheet['DA20'] = f'{formatted_date}'
    else:
        sheet['DA20'] = ''

    if data['period_by']:
        date_str = str(data['period_by'])
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        formatted_date = date_obj.strftime("%d-%m-%Y")
        sheet['DN20'] = f'{formatted_date}'

    else:
        sheet['DN20'] = ''

    if data["price_outlay"]:
        sheet['BX24'] = f'{data["price_outlay"]}'
    else:
        sheet['BX24'] = '0'

    if data['nds'] == -1 or data['nds'] == '-1':
        sheet['ET27'] = 'стоимость, руб (Без НДС).'
    else:
        sheet['ET27'] = f'стоимость, руб (НДС {data["nds"]}%).'

    start_table_row = 29
    total_sum = 0
    total_q = 0
    if int(data['nds']) > 0 and int(data['nds']):
        nds = int(data['nds'])
    else:
        nds = 0

    for idx, table_data in enumerate(formset_data, 1):
        total_q += table_data['quantity']

        sheet[f'A{start_table_row + idx}'] = f'{idx}'
        sheet[f'I{start_table_row + idx}'] = table_data['number_outlay']
        sheet[f'R{start_table_row + idx}'] = table_data['name']
        sheet[f'CP{start_table_row + idx}'] = table_data['number_unit']
        sheet[f'DB{start_table_row + idx}'] = table_data['unit_of_measurement']
        sheet[f'DN{start_table_row + idx}'] = f"{table_data['quantity']}"
        sheet[f'ED{start_table_row + idx}'] = table_data['price']
        sheet[f'ET{start_table_row + idx}'] = table_data['amount']
        total_sum += table_data['amount']

    sheet[f'DN{start_table_row + len(formset_data) + 1}'] = f'{total_q}'
    sheet[f'ET{start_table_row + len(formset_data) + 1}'] = f'{total_sum}'

    sheet[f'Q{start_table_row + len(formset_data) + 4}'] = f'{data["organization"].position_at_work}'
    sheet[f'CK{start_table_row + len(formset_data) + 4}'] = f'{data["organization"].supervisor}'

    sheet[f'Q{start_table_row + len(formset_data) + 8}'] = ''
    sheet[f'CK{start_table_row + len(formset_data) + 8}'] = ''

    if data['organization'].stamp and data['is_stamp']:
        image_file = data['organization'].stamp
        img = Image(BytesIO(image_file.read()))

        img.width = 45 * 2.83
        img.height = 45 * 2.83

        sheet.add_image(img, f"S{start_table_row + len(formset_data) + 2}")

    if data['organization'].signature and data['is_stamp']:
        image_file = data['organization'].signature

        image_data = image_file.read()

        img_stream = BytesIO(image_data)

        img1 = Image(img_stream)
        img1.width = 80
        img1.height = 30
        sheet.add_image(img1, f"BC{start_table_row + len(formset_data) + 4}")

    if pdf:
        convertapi.api_credentials = 'secret_wBUU4YjxTpfeIPwA'

        temp_excel_path = "ks_2/utils/invoice.xlsx"
        temp_modified_pdf_path = "ks_2/utils/invoice_modified.pdf"

        workbook.save(temp_excel_path)

        temp_pdf_path = convertapi.convert('pdf', {
            'File': temp_excel_path,
        }, from_format='xls').save_files('ks_2/utils')[0]

        reader = PdfReader(temp_pdf_path)
        writer = PdfWriter()

        pages_to_remove = [i for i in range(1, 55)]

        for i in range(len(reader.pages)):
            writer.add_page(reader.pages[i])
            # if i not in pages_to_remove:
            #     writer.add_page(reader.pages[i])

        with open(temp_modified_pdf_path, "wb") as output_pdf:
            writer.write(output_pdf)

        with open(temp_modified_pdf_path, "rb") as pdf_file:
            response = HttpResponse(pdf_file.read(), content_type="application/pdf")
            if watch_document:
                response["Content-Disposition"] = "inline; filename=ks_2.pdf"
            else:
                response["Content-Disposition"] = "attachment; filename=ks_2.pdf"

        os.remove(temp_excel_path)
        os.remove(temp_pdf_path)
        os.remove(temp_modified_pdf_path)

        return response

        # temp_excel_path = "ks_2/utils/invoice.xlsx"
        # temp_pdf_path = "ks_2/utils/invoice.pdf"
        # temp_modified_pdf_path = "ks_2/utils/invoice_modified.pdf"
        #
        # workbook.save(temp_excel_path)
        #
        # workbook_aspose = Workbook(temp_excel_path)
        # workbook_aspose.save(temp_pdf_path, SaveFormat.PDF)
        #
        # reader = PdfReader(temp_pdf_path)
        # writer = PdfWriter()
        #
        # pages_to_remove = [i for i in range(1, 55)]
        #
        # for i in range(len(reader.pages)):
        #     if i not in pages_to_remove:
        #         writer.add_page(reader.pages[i])
        #
        # with open(temp_modified_pdf_path, "wb") as output_pdf:
        #     writer.write(output_pdf)
        #
        # with open(temp_modified_pdf_path, "rb") as pdf_file:
        #     response = HttpResponse(pdf_file.read(), content_type="application/pdf")
        #     if watch_document:
        #         response["Content-Disposition"] = "inline; filename=invoice.pdf"
        #     else:
        #         response["Content-Disposition"] = "attachment; filename=invoice.pdf"
        #
        # os.remove(temp_excel_path)
        # os.remove(temp_pdf_path)
        # os.remove(temp_modified_pdf_path)
        #
        # return response

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=ks_2.xlsx"
    workbook.save(response)
    return response
