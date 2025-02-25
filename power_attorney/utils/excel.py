import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.worksheet.table import Table, TableStyleInfo
from aspose.cells import Workbook, HtmlSaveOptions, SaveFormat, License
from bs4 import BeautifulSoup
import aspose.cells as cells
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime
from PyPDF2 import PdfReader, PdfWriter
import os

license_as = License()
license_as.set_license("lic/Aspose.TotalforPythonvia.NET.lic")

months_russian = [
    'Января', 'Февраля', 'Марта', 'Апреля', 'Мая', 'Июня',
    'Июля', 'Августа', 'Сентября', 'Октября', 'Ноября', 'Декабря'
]


def html_to_excel():
    html_file = "power_attorney/utils/updated_file.html"

    html_load_options = cells.HtmlLoadOptions()

    workbook = cells.Workbook(html_file, html_load_options)

    workbook.save("power_attorney/utils/output.xlsx")


def change_html(count_rows):
    with open("power_attorney/utils/Счет_на_оплату.html", "r",
              encoding="utf-8") as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, "html.parser")

    element_to_remove = soup.find("div",
                                  string=lambda text: text and "Evaluation Only. Created with Aspose.Cells" in text)

    if element_to_remove:
        element_to_remove.decompose()

    target_row = soup.find("td", string=lambda
        text: text and "Beton" in text)

    if target_row:
        parent_row = target_row.find_parent("tr")

        if parent_row:
            num_rows_to_add = count_rows - 1

            if num_rows_to_add >= 1:
                for _ in range(num_rows_to_add):
                    new_row = BeautifulSoup(str(parent_row), "html.parser")

                    parent_row.insert_after(new_row)

            with open("power_attorney/utils/updated_file.html",
                      "w",
                      encoding="utf-8") as file:
                file.write(str(soup))


def excel_to_html():
    file_path = "power_attorney/utils/Доверенность №1 от 05.02.2025 г..xlsx"
    workbook = Workbook(file_path)

    save_options = HtmlSaveOptions()
    save_options.export_active_worksheet_only = True
    save_options.export_images_as_base64 = True

    html_file = "power_attorney/utils/Счет_на_оплату.html"
    workbook.save(html_file, save_options)


def create_power_attorney_excel(data, formset_data, pdf=False, watch_document=False):
    excel_to_html()
    change_html(len(formset_data))
    html_to_excel()

    file_path = 'power_attorney/utils/output.xlsx'

    workbook = openpyxl.load_workbook(file_path)
    workbook.remove(workbook["Evaluation Warning"])

    sheet = workbook["Доверенность"]

    for col in sheet.columns:
        try:
            sheet.column_dimensions[col[0].column_letter].width = 1
        except:
            continue

    for row in sheet.iter_rows():
        if row[0].row in [19, 21, 29]:
            sheet.row_dimensions[row[0].row].height = 22
        elif row[0].row == 9:
            sheet.row_dimensions[row[0].row].height = 27

    sheet['A3'] = f'{data["name"]}'

    date_str = str(data['date'])
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d-%m-%Y")
    sheet['O3'] = f'{formatted_date}'

    if data['validity_period']:
        date_str = str(data['validity_period'])
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        formatted_date = date_obj.strftime("%d-%m-%Y")
        sheet['Z3'] = f'{formatted_date}'
    else:
        sheet['Z3'] = ''

    if data["person_power"]:
        sheet['AK3'] = f'{data["person_power"]}'
    else:
        sheet['AK3'] = ''

    if data["to_receive_from"]:
        sheet['A5'] = f'{data["to_receive_from"]}'
    else:
        sheet['A5'] = ''

    if data["according_document"]:
        sheet['AS5'] = f'{data["according_document"]}'
    else:
        sheet['AS5'] = ''

    sheet[
        'M15'] = f'{data["organization"].naming}, ИНН {data["organization"].inn} КПП {data["organization"].kpp}, {data["organization"].address}'
    sheet['A17'] = f'Доверенность № {data["name"]}'

    date_str = str(data['date'])
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d %B %Y").split(' ')

    month = date_obj.month
    month_russian = months_russian[month - 1]

    sheet['O19'] = f'{formatted_date[0]}'
    sheet['U19'] = f'{month_russian}'
    sheet['AN19'] = f'{formatted_date[2]}'

    if data['validity_period']:
        date_str = str(data['validity_period'])
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        formatted_date = date_obj.strftime("%d %B %Y").split(' ')

        month = date_obj.month
        month_russian = months_russian[month - 1]

        sheet['AF21'] = f'{formatted_date[0]}'
        sheet['AL21'] = f'{month_russian}'
        sheet['BE21'] = f'{formatted_date[2]}'
    else:
        sheet['AF21'] = ''
        sheet['AL21'] = ''
        sheet['BE21'] = ''

    sheet[
        'A23'] = f'{data["organization"].naming}, ИНН {data["organization"].inn} КПП {data["organization"].kpp}, {data["organization"].address}'
    sheet[
        'A26'] = f'{data["organization"].naming}, ИНН {data["organization"].inn} КПП {data["organization"].kpp}, {data["organization"].address}'

    if data["bank_organization"]:
        sheet[
            'A29'] = f'Счет № {data["bank_organization"].current_account} в {data["bank_organization"].naming}, {data["bank_organization"].location}, БИК {data["bank_organization"].bic}, корр.сч. {data["bank_organization"].correspondent_account}'
    else:
        sheet[
            'A29'] = ''

    if data["person_power"]:
        sheet['A31'] = f'Доверенность выдана: {data["person_power"]}'
    else:
        sheet['A31'] = f'Доверенность выдана:'

    if data["passport_series"] and data["passport_number"]:
        sheet['A33'] = f'Паспорт: {data["passport_series"]} {data["passport_number"]}'
    else:
        sheet['A33'] = 'Паспорт:'

    if data["issued_by"]:
        sheet['A35'] = f'Кем выдан: {data["issued_by"]}'
    else:
        sheet['A35'] = 'Кем выдан:'

    if data['date_issue']:
        date_str = str(data['date_issue'])
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        formatted_date = date_obj.strftime("%d-%m-%Y")
        sheet['A37'] = f'Дата выдачи: {formatted_date}'
    else:
        sheet['A37'] = f'Дата выдачи:'

    if data["to_receive_from"]:
        sheet['A39'] = f'На получение от {data["to_receive_from"]}'
    else:
        sheet['A39'] = f'На получение от'

    if data["according_document"]:
        sheet['A41'] = f'материальных ценностей по {data["according_document"]}'
    else:
        sheet['A41'] = f'материальных ценностей по'

    start_table_row = 45

    for idx, table_data in enumerate(formset_data, 1):
        sheet[f'A{start_table_row + idx}'] = f'{idx}'
        sheet[f'E{start_table_row + idx}'] = table_data['name']
        sheet[f'CG{start_table_row + idx}'] = table_data['unit_of_measurement']
        sheet[f'CN{start_table_row + idx}'] = f"{table_data['quantity']}"

    sheet[f'BA{start_table_row + len(formset_data) + 4}'] = f'{data["organization"].supervisor}'

    if data["organization"].accountant:
        sheet[f'BA{start_table_row + len(formset_data) + 8}'] = f'{data["organization"].accountant}'
    else:
        sheet[f'BA{start_table_row + len(formset_data) + 8}'] = ''

    if data['organization'].stamp and data['is_stamp']:
        image_file = data['organization'].stamp
        img = Image(BytesIO(image_file.read()))

        img.width = 45 * 2.83
        img.height = 45 * 2.83

        sheet.add_image(img, f"P{start_table_row + len(formset_data) + 3}")

    if data['organization'].signature and data['is_stamp']:
        image_file = data['organization'].signature

        image_data = image_file.read()

        img_stream = BytesIO(image_data)

        img1 = Image(img_stream)
        img1.width = 70
        img1.height = 25
        sheet.add_image(img1, f"Z{start_table_row + len(formset_data) + 4}")

    for row in sheet.iter_rows():
        if row[0].row in [start_table_row + len(formset_data) + 2, start_table_row + len(formset_data) + 4]:
            sheet.row_dimensions[row[0].row].height = 22

    if pdf:
        temp_excel_path = "power_attorney/utils/invoice.xlsx"
        temp_pdf_path = "power_attorney/utils/invoice.pdf"
        temp_modified_pdf_path = "power_attorney/utils/invoice_modified.pdf"

        workbook.save(temp_excel_path)

        workbook_aspose = Workbook(temp_excel_path)
        workbook_aspose.save(temp_pdf_path, SaveFormat.PDF)

        reader = PdfReader(temp_pdf_path)
        writer = PdfWriter()

        pages_to_remove = [i for i in range(1, 55)]

        for i in range(len(reader.pages)):
            if i not in pages_to_remove:
                writer.add_page(reader.pages[i])

        with open(temp_modified_pdf_path, "wb") as output_pdf:
            writer.write(output_pdf)

        with open(temp_modified_pdf_path, "rb") as pdf_file:
            response = HttpResponse(pdf_file.read(), content_type="application/pdf")
            if watch_document:
                response["Content-Disposition"] = "inline; filename=invoice.pdf"
            else:
                response["Content-Disposition"] = "attachment; filename=invoice.pdf"

        os.remove(temp_excel_path)
        os.remove(temp_pdf_path)
        os.remove(temp_modified_pdf_path)

        return response

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=invoice.xlsx"
    workbook.save(response)
    return response
