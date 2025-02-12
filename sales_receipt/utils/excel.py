import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.worksheet.table import Table, TableStyleInfo
from aspose.cells import Workbook, HtmlSaveOptions, SaveFormat
from bs4 import BeautifulSoup
import aspose.cells as cells
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime
from PyPDF2 import PdfReader, PdfWriter
import os


def html_to_excel():
    html_file = "sales_receipt/utils/updated_file.html"

    html_load_options = cells.HtmlLoadOptions()

    workbook = cells.Workbook(html_file, html_load_options)

    workbook.save("sales_receipt/utils/output.xlsx")


def change_html(count_rows):
    with open("sales_receipt/utils/Счет_на_оплату.html", "r",
              encoding="utf-8") as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, "html.parser")

    element_to_remove = soup.find("div",
                                  string=lambda text: text and "Evaluation Only. Created with Aspose.Cells" in text)

    if element_to_remove:
        element_to_remove.decompose()

    target_row = soup.find("td", string=lambda
        text: text and "Молоко" in text)

    if target_row:
        parent_row = target_row.find_parent("tr")

        if parent_row:
            num_rows_to_add = count_rows - 1

            if num_rows_to_add >= 1:
                for _ in range(num_rows_to_add):
                    new_row = BeautifulSoup(str(parent_row), "html.parser")

                    parent_row.insert_after(new_row)

            with open("sales_receipt/utils/updated_file.html",
                      "w",
                      encoding="utf-8") as file:
                file.write(str(soup))


def excel_to_html():
    file_path = "sales_receipt/utils/Товарный чек №1 от 06.02.2025 г..xlsx"
    workbook = Workbook(file_path)

    save_options = HtmlSaveOptions()
    save_options.export_active_worksheet_only = True
    save_options.export_images_as_base64 = True

    html_file = "sales_receipt/utils/Счет_на_оплату.html"
    workbook.save(html_file, save_options)


def create_sales_receipt_excel(data, formset_data, pdf=False):
    excel_to_html()
    change_html(len(formset_data))
    html_to_excel()

    file_path = 'sales_receipt/utils/output.xlsx'

    workbook = openpyxl.load_workbook(file_path)
    workbook.remove(workbook["Evaluation Warning"])

    sheet = workbook["Товарный чек"]

    for col in sheet.columns:
        try:
            sheet.column_dimensions[col[0].column_letter].width = 1
        except:
            continue

    sheet['A1'] = f'{data["organization"].naming}, ИНН {data["organization"].inn}'
    sheet['A2'] = f'{data["organization"].address}'

    date_str = str(data['date'])
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d-%m-%Y")
    sheet['A4'] = f'Товарный чек № {data["name"]} от {formatted_date}'

    start_table_row = 6
    total_sum = 0

    for idx, table_data in enumerate(formset_data, 1):
        total_sum += table_data["amount"]

        sheet[f'A{start_table_row + idx}'] = f'{idx}'
        sheet[f'E{start_table_row + idx}'] = f'{table_data["article_number"]}'
        sheet[f'N{start_table_row + idx}'] = table_data['name']
        sheet[f'BL{start_table_row + idx}'] = table_data['unit_of_measurement']
        sheet[f'BS{start_table_row + idx}'] = f'{table_data["quantity"]}'
        sheet[f'CB{start_table_row + idx}'] = f'{table_data["price"]}'
        sheet[f'CQ{start_table_row + idx}'] = f'{table_data["amount"]}'

    sheet[f'CQ{start_table_row + len(formset_data) + 1}'] = f'{total_sum}'

    sheet[
        f'A{start_table_row + len(formset_data) + 3}'] = f'Всего наименований {len(formset_data)}, на сумму {total_sum} руб.'
    sheet[f'A{start_table_row + len(formset_data) + 4}'] = ''

    sheet[f'AI{start_table_row + len(formset_data) + 6}'] = data["organization"].position_at_work
    sheet[f'BO{start_table_row + len(formset_data) + 6}'] = data["organization"].supervisor

    if data['organization'].stamp and data['is_stamp']:
        image_file = data['organization'].stamp
        img = Image(BytesIO(image_file.read()))

        img.width = 45 * 2.83
        img.height = 45 * 2.83

        sheet.add_image(img, f"BO{start_table_row + len(formset_data) + 5}")

    if data['organization'].signature and data['is_stamp']:
        image_file = data['organization'].signature

        image_data = image_file.read()

        img_stream = BytesIO(image_data)

        img1 = Image(img_stream)
        img1.width = 70
        img1.height = 25
        sheet.add_image(img1, f"J{start_table_row + len(formset_data) + 6}")

    if pdf:
        temp_excel_path = "sales_receipt/utils/invoice.xlsx"
        temp_pdf_path = "sales_receipt/utils/invoice.pdf"
        temp_modified_pdf_path = "sales_receipt/utils/invoice_modified.pdf"

        workbook.save(temp_excel_path)

        workbook_aspose = Workbook(temp_excel_path)
        workbook_aspose.save(temp_pdf_path, SaveFormat.PDF)

        reader = PdfReader(temp_pdf_path)
        writer = PdfWriter()

        pages_to_remove = [i for i in range(1, 55)]

        for i in range(len(reader.pages)):
            if i not in pages_to_remove:
                writer.add_page(reader.pages[i])

        with open(temp_modified_pdf_path, "wb") as output_pdf:
            writer.write(output_pdf)

        with open(temp_modified_pdf_path, "rb") as pdf_file:
            response = HttpResponse(pdf_file.read(), content_type="application/pdf")
            response["Content-Disposition"] = "attachment; filename=invoice.pdf"

        os.remove(temp_excel_path)
        os.remove(temp_pdf_path)
        os.remove(temp_modified_pdf_path)

        return response

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=invoice.xlsx"
    workbook.save(response)
    return response
