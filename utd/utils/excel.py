import convertapi
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.worksheet.table import Table, TableStyleInfo
from aspose.cells import Workbook, HtmlSaveOptions, SaveFormat, License
from bs4 import BeautifulSoup
import aspose.cells as cells
from datetime import datetime
from openpyxl.drawing.image import Image
from io import BytesIO
import locale
from PyPDF2 import PdfReader, PdfWriter
import os
import uuid
import math
import re
from openpyxl.worksheet.pagebreak import Break
import pdfplumber

months_russian = [
    'Января', 'Февраля', 'Марта', 'Апреля', 'Мая', 'Июня',
    'Июля', 'Августа', 'Сентября', 'Октября', 'Ноября', 'Декабря'
]


def html_to_excel():
    html_file = "utd/utils/updated_file.html"

    html_load_options = cells.HtmlLoadOptions()

    workbook = cells.Workbook(html_file, html_load_options)

    workbook.save("utd/utils/output.xlsx")


def change_html(count_rows):
    with open("utd/utils/Счет_на_оплату.html", "r",
              encoding="utf-8") as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, "html.parser")

    element_to_remove = soup.find("div",
                                  string=lambda text: text and "Evaluation Only. Created with Aspose.Cells" in text)

    if element_to_remove:
        element_to_remove.decompose()

    target_row = soup.find("td", string=lambda
        text: text and "MacBook Air" in text)

    if target_row:
        parent_row = target_row.find_parent("tr")

        if parent_row:
            num_rows_to_add = count_rows - 1

            if num_rows_to_add >= 1:
                for _ in range(num_rows_to_add):
                    new_row = BeautifulSoup(str(parent_row), "html.parser")

                    parent_row.insert_after(new_row)

            with open("utd/utils/updated_file.html", "w",
                      encoding="utf-8") as file:
                file.write(str(soup))


def excel_to_html():
    file_path = "utd/utils/Универсальный передаточный документ №3 от 31.01.2025 г..xlsx"
    workbook = Workbook(file_path)

    save_options = HtmlSaveOptions()
    save_options.export_active_worksheet_only = True
    save_options.export_images_as_base64 = True

    html_file = "utd/utils/Счет_на_оплату.html"
    workbook.save(html_file, save_options)


def create_utd_excel(data, formset_data, pdf=False, watch_document=False):
    excel_to_html()
    change_html(len(formset_data))
    html_to_excel()

    file_path = 'utd/utils/output.xlsx'

    workbook = openpyxl.load_workbook(file_path)
    workbook.remove(workbook["Evaluation Warning"])

    sheet = workbook["УПД"]

    # for row in sheet.iter_rows():
    #     if row[0].row in [7, 9, 10, 11, 12]:
    #         sheet.row_dimensions[row[0].row].height = 22
    # if len(formset_data) > 1:
    #     for row_idx in range(42, sheet.max_row, 42):
    #         sheet.row_breaks.append(Break(id=row_idx))

    start_col = 208
    num_cols = 790

    sheet.delete_cols(start_col, num_cols)

    sheet['O2'] = 'УПД'
    sheet['AP2'] = data['name']

    if data['type_document'] == 'Счет-фактура и передаточный документ(акт)':
        sheet['H7'] = '1'
    else:
        sheet['H7'] = '2'

    date_str = str(data['date'])
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d %B %Y").split(' ')

    month = date_obj.month
    month_russian = months_russian[month - 1]

    sheet['BK2'] = f'{formatted_date[0]} {month_russian} {formatted_date[2]}'

    sheet['AP6'] = data['organization'].naming
    sheet['AP7'] = data['organization'].address
    if data["organization"].kpp:
        sheet['AP8'] = f'{data["organization"].inn} / {data["organization"].kpp}'
    else:
        sheet['AP8'] = f'{data["organization"].inn}'
    if data["shipper"]:
        if data["organization"].id == data["shipper"].id:
            sheet['AP9'] = f'он же'
        else:
            sheet['AP9'] = f'{data["shipper"].naming}, {data["shipper"].address}'
    else:
        sheet['AP9'] = ''
    if data["consignee"]:
        if data["counterparty"].id == data["consignee"].id:
            sheet['AP10'] = f'он же'
        else:
            sheet['AP10'] = f'{data["consignee"].naming}, {data["consignee"].address}'
    else:
        sheet['AP10'] = ''
    sheet['AT11'] = data['payment_document']
    sheet['AX12'] = data['shipping_document']

    if data["counterparty"]:
        sheet['DL6'] = f'{data["counterparty"].naming}'
        sheet['DL7'] = f'{data["counterparty"].address}'
        if data["counterparty"].kpp:
            sheet['DL8'] = f'{data["counterparty"].inn} / {data["counterparty"].kpp}'
        else:
            sheet['DL8'] = f'{data["counterparty"].inn}'
    else:
        sheet['DL6'] = ''
        sheet['DL7'] = ''
        sheet['DL8'] = ''
    sheet['EE10'] = data['state_ID_contract']

    start_table_row = 21
    total_sum = 0
    total_sum_nds = 0
    total_sum_with_nds = 0

    max_symbol_line = 25
    max_symbol_line_address = 35

    height_line = 9
    height_line_table = 10

    if int(data['nds']) > 0:
        nds = int(data['nds'])
    else:
        nds = 0

    if data['organization']:
        len_address_org = math.ceil(len(data['organization'].address) / max_symbol_line_address)
        len_address_coun = math.ceil(len(data['counterparty'].address) / max_symbol_line_address)

        len_address = max(len_address_org, len_address_coun)

        if len_address != 0:
            sheet.row_dimensions[7].height = len_address * height_line

    if data["shipper"]:
        if sheet['AP9'].value == 'он же' or sheet['AP9'].value == '':
            sheet.row_dimensions[9].height = 12
        else:
            len_shipper = math.ceil((len(data["shipper"].naming) + len(data["shipper"].address)) / max_symbol_line)
            if len_shipper != 0:
                sheet.row_dimensions[9].height = len_shipper * height_line

    if data["consignee"]:
        if sheet['AP10'].value == 'он же' or sheet['AP10'].value == '':
            sheet.row_dimensions[9].height = 12
        else:
            len_consignee = math.ceil((len(data["consignee"].naming) + len(data["consignee"].address)) / max_symbol_line)
            if len_consignee != 0:
                sheet.row_dimensions[10].height = len_consignee * height_line

    for idx, table_data in enumerate(formset_data, 1):
        # sheet.row_dimensions[start_table_row + idx].height = 20
        len_name = math.ceil(len(table_data['name']) / max_symbol_line)

        if len_name != 0:
            sheet.row_dimensions[start_table_row + idx].height = len_name * height_line_table

        total_sum += table_data["amount"]

        if table_data["product_code"]:
            sheet[f'A{start_table_row + idx}'] = f'{table_data["product_code"]}'
        else:
            sheet[f'A{start_table_row + idx}'] = ''
        sheet[f'N{start_table_row + idx}'] = f'{idx}'
        sheet[f'R{start_table_row + idx}'] = f'{table_data["name"]}'

        if table_data["product_type_code"]:
            sheet[f'AQ{start_table_row + idx}'] = f'{table_data["product_type_code"]}'
        else:
            sheet[f'AQ{start_table_row + idx}'] = ''

        sheet[f'AW{start_table_row + idx}'] = ''
        sheet[f'BA{start_table_row + idx}'] = f'{table_data["unit_of_measurement"]}'
        sheet[f'BN{start_table_row + idx}'] = f'{table_data["quantity"]}'
        sheet[f'BU{start_table_row + idx}'] = f'{table_data["price"]}'
        sheet[f'CE{start_table_row + idx}'] = f'{round(float(table_data["quantity"]) * float(table_data["price"]), 2)}'
        total_sum_with_nds += float(table_data["quantity"]) * float(table_data["price"])

        if table_data["excise"]:
            sheet[f'CR{start_table_row + idx}'] = f'{table_data["excise"]}'
        else:
            sheet[f'CR{start_table_row + idx}'] = 'Без акциза'
        if data["nds"] == '-1' or data["nds"] == -1:
            sheet[f'CX{start_table_row + idx}'] = 'Без НДС'
        else:
            sheet[f'CX{start_table_row + idx}'] = f'{data["nds"]}%'
        if nds > 0:
            sheet[
                f'DD{start_table_row + idx}'] = f'{round(float(table_data["quantity"]) * float(table_data["price"]) * nds * 0.01, 2)}'
            total_sum_nds += float(table_data["quantity"]) * float(table_data["price"]) * nds * 0.01
        else:
            sheet[f'DD{start_table_row + idx}'] = f'{0}'
        sheet[
            f'DQ{start_table_row + idx}'] = f'{table_data["amount"]}'

        if table_data["country"]:
            sheet[f'ED{start_table_row + idx}'] = ''
            sheet[f'EJ{start_table_row + idx}'] = f'{table_data["country"]}'
        else:
            sheet[f'ED{start_table_row + idx}'] = ''
            sheet[f'EJ{start_table_row + idx}'] = ''

        if table_data["number_GTD"]:
            sheet[f'ET{start_table_row + idx}'] = f'{table_data["number_GTD"]}'
        else:
            sheet[f'ET{start_table_row + idx}'] = ''

    sheet[f'DD{start_table_row + len(formset_data) + 1}'] = f'{round(total_sum_nds, 2)}'
    sheet[f'CE{start_table_row + len(formset_data) + 1}'] = f'{round(total_sum_with_nds, 2)}'
    sheet[f'DQ{start_table_row + len(formset_data) + 1}'] = f'{float(total_sum)}'

    if data['organization']:
        sheet[f'BS{start_table_row + len(formset_data) + 3}'] = data['organization'].supervisor
        sheet[f'EL{start_table_row + len(formset_data) + 3}'] = data['organization'].accountant
    else:
        sheet[f'BS{start_table_row + len(formset_data) + 3}'] = ''
        sheet[f'EL{start_table_row + len(formset_data) + 3}'] = ''

    sheet[f'AW{start_table_row + len(formset_data) + 9}'] = data['basis_for_transfer']
    sheet[f'AJ{start_table_row + len(formset_data) + 11}'] = data['data_transportation']

    if data['organization']:
        sheet[f'A{start_table_row + len(formset_data) + 14}'] = data['organization'].position_at_work
        sheet[f'AZ{start_table_row + len(formset_data) + 14}'] = data['organization'].supervisor
    else:
        sheet[f'A{start_table_row + len(formset_data) + 14}'] = ''
        sheet[f'AZ{start_table_row + len(formset_data) + 14}'] = ''

    if data['shipment_date']:
        date_str = str(data['shipment_date'])
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        formatted_date = date_obj.strftime("%d %B %Y").split(' ')

        month = date_obj.month
        month_russian = months_russian[month - 1]

        sheet[f'AL{start_table_row + len(formset_data) + 16}'] = formatted_date[0]
        sheet[f'AR{start_table_row + len(formset_data) + 16}'] = month_russian
        sheet[f'BP{start_table_row + len(formset_data) + 16}'] = formatted_date[2]
    else:
        sheet[f'AL{start_table_row + len(formset_data) + 16}'] = ''
        sheet[f'AR{start_table_row + len(formset_data) + 16}'] = ''
        sheet[f'BP{start_table_row + len(formset_data) + 16}'] = ''

    sheet[f'A{start_table_row + len(formset_data) + 21}'] = data['organization'].position_at_work
    sheet[f'AZ{start_table_row + len(formset_data) + 21}'] = data['organization'].supervisor

    sheet[
        f'A{start_table_row + len(formset_data) + 24}'] = ''

    sheet[f'CF{start_table_row + len(formset_data) + 14}'] = ''
    sheet[f'EE{start_table_row + len(formset_data) + 14}'] = ''

    if data['date_of_receipt']:
        date_str = str(data['date_of_receipt'])
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        formatted_date = date_obj.strftime("%d %B %Y").split(' ')

        month = date_obj.month
        month_russian = months_russian[month - 1]

        sheet[f'DP{start_table_row + len(formset_data) + 16}'] = formatted_date[0]
        sheet[f'DV{start_table_row + len(formset_data) + 16}'] = month_russian
        sheet[f'ET{start_table_row + len(formset_data) + 16}'] = formatted_date[2]
    else:
        sheet[f'DP{start_table_row + len(formset_data) + 16}'] = ''
        sheet[f'DV{start_table_row + len(formset_data) + 16}'] = ''
        sheet[f'ET{start_table_row + len(formset_data) + 16}'] = ''

    sheet[f'CF{start_table_row + len(formset_data) + 21}'] = ''
    sheet[f'EE{start_table_row + len(formset_data) + 21}'] = ''

    sheet[
        f'CF{start_table_row + len(formset_data) + 24}'] = ''

    if data['organization'].stamp and data['is_stamp']:
        image_file = data['organization'].stamp

        image_data = image_file.read()

        img_stream1 = BytesIO(image_data)
        img1 = Image(img_stream1)
        img1.width = 43 * 2.83
        img1.height = 43 * 2.83

        if len(formset_data) == 1:
            sheet.add_image(img1, f"A{start_table_row + len(formset_data) + 20}")
        else:
            sheet.add_image(img1, f"A{start_table_row + len(formset_data) + 20}")

    if data['organization'].signature and data['is_stamp']:
        image_file = data['organization'].signature

        image_data = image_file.read()

        img_stream = BytesIO(image_data)

        img1 = Image(img_stream)
        img1.width = 70
        img1.height = 37
        sheet.add_image(img1, f"AD{start_table_row + len(formset_data) + 14}")

        img_stream2 = BytesIO(image_data)
        img2 = Image(img_stream2)
        img2.width = 70
        img2.height = 37
        sheet.add_image(img2, f"AD{start_table_row + len(formset_data) + 21}")

        img_stream3 = BytesIO(image_data)
        img3 = Image(img_stream3)
        img3.width = 70
        img3.height = 37
        sheet.add_image(img3, f"AZ{start_table_row + len(formset_data) + 3}")

    for row in sheet.iter_rows():
        if row[0].row in [start_table_row + len(formset_data) + 3, start_table_row + len(formset_data) + 14,
                          start_table_row + len(formset_data) + 21]:
            sheet.row_dimensions[row[0].row].height = 20

    if pdf:
        convertapi.api_credentials = 'secret_omNCSVvj1fl5oFYe'

        temp_excel_path = "utd/utils/invoice.xlsx"
        temp_modified_pdf_path = "utd/utils/invoice_modified.pdf"
        temp_excel_count = f"utd/utils/count_page_excel_{uuid.uuid4().hex}.xlsx"

        workbook.save(temp_excel_count)
        workbook.close()

        if len(formset_data) > 1:
            temp_pdf_path_count = convertapi.convert('pdf', {
                'File': temp_excel_count,
                'Scale': '100'
            }, from_format='xls').save_files('utd/utils')[0]
        else:
            temp_pdf_path_count = convertapi.convert('pdf', {
                'File': temp_excel_count,
                'Scale': '93'
            }, from_format='xls').save_files('utd/utils')[0]

        with pdfplumber.open(temp_pdf_path_count) as pdf:
            last_page = pdf.pages[-1]
            text = last_page.extract_text()
            if text:
                lines = text.split('\n')
                line_count = len(lines)
            else:
                line_count = 0

        if line_count == 1:
            workbook_cells = Workbook(temp_excel_count)
            worksheet_cells = workbook_cells.worksheets[0]

            worksheet_cells.cells.insert_rows(0, 7)

            workbook_cells.save(temp_excel_count)
        elif line_count == 2:
            workbook_cells = Workbook(temp_excel_count)
            worksheet_cells = workbook_cells.worksheets[0]

            worksheet_cells.cells.insert_rows(0, 6)

            workbook_cells.save(temp_excel_count)
        elif line_count == 3:
            workbook_cells = Workbook(temp_excel_count)
            worksheet_cells = workbook_cells.worksheets[0]

            worksheet_cells.cells.insert_rows(0, 5)

            workbook_cells.save(temp_excel_count)
        elif line_count == 4:
            workbook_cells = Workbook(temp_excel_count)
            worksheet_cells = workbook_cells.worksheets[0]

            worksheet_cells.cells.insert_rows(0, 4)

            workbook_cells.save(temp_excel_count)
        elif line_count == 5:
            workbook_cells = Workbook(temp_excel_count)
            worksheet_cells = workbook_cells.worksheets[0]

            worksheet_cells.cells.insert_rows(0, 3)

            workbook_cells.save(temp_excel_count)
        elif line_count == 6:
            workbook_cells = Workbook(temp_excel_count)
            worksheet_cells = workbook_cells.worksheets[0]

            worksheet_cells.cells.insert_rows(0, 2)

            workbook_cells.save(temp_excel_count)
        elif line_count == 7:
            workbook_cells = Workbook(temp_excel_count)
            worksheet_cells = workbook_cells.worksheets[0]

            worksheet_cells.cells.insert_rows(0, 1)

            workbook_cells.save(temp_excel_count)

        reader_count = PdfReader(temp_pdf_path_count)
        pages_count = len(reader_count.pages)

        workbook = openpyxl.load_workbook(temp_excel_count)
        sheet = workbook["УПД"]
        # if len(formset_data) == 1:
        #     sheet[f"A{start_table_row + len(formset_data) + 6}"] = f'1'
        # else:
        if line_count == 1:
            sheet[f"A{start_table_row + 7 + len(formset_data) + 6}"] = f'{pages_count}'
            workbook.remove(workbook["Evaluation Warning"])
        elif line_count == 2:
            sheet[f"A{start_table_row + 6 + len(formset_data) + 6}"] = f'{pages_count}'
            workbook.remove(workbook["Evaluation Warning"])
        elif line_count == 3:
            sheet[f"A{start_table_row + 5 + len(formset_data) + 6}"] = f'{pages_count}'
            workbook.remove(workbook["Evaluation Warning"])
        elif line_count == 4:
            sheet[f"A{start_table_row + 4 + len(formset_data) + 6}"] = f'{pages_count}'
            workbook.remove(workbook["Evaluation Warning"])
        elif line_count == 5:
            sheet[f"A{start_table_row + 3 + len(formset_data) + 6}"] = f'{pages_count}'
            workbook.remove(workbook["Evaluation Warning"])
        elif line_count == 6:
            sheet[f"A{start_table_row + 2 + len(formset_data) + 6}"] = f'{pages_count}'
            workbook.remove(workbook["Evaluation Warning"])
        elif line_count == 7:
            sheet[f"A{start_table_row + 1 + len(formset_data) + 6}"] = f'{pages_count}'
            workbook.remove(workbook["Evaluation Warning"])
        else:
            sheet[f"A{start_table_row + len(formset_data) + 6}"] = f'{pages_count}'

        workbook.save(temp_excel_path)
        workbook.close()

        if len(formset_data) > 1:
            temp_pdf_path = convertapi.convert('pdf', {
                'File': temp_excel_path,
                'Scale': '100'
            }, from_format='xls').save_files('utd/utils')[0]
        else:
            temp_pdf_path = convertapi.convert('pdf', {
                'File': temp_excel_path,
                'Scale': '93'
            }, from_format='xls').save_files('utd/utils')[0]

        reader = PdfReader(temp_pdf_path)
        writer = PdfWriter()

        pages_to_remove = [i for i in range(1, 55)]

        for i in range(len(reader.pages)):
            writer.add_page(reader.pages[i])
            # if i not in pages_to_remove:
            #     writer.add_page(reader.pages[i])

        with open(temp_modified_pdf_path, "wb") as output_pdf:
            writer.write(output_pdf)

        with open(temp_modified_pdf_path, "rb") as pdf_file:
            response = HttpResponse(pdf_file.read(), content_type="application/pdf")
            if watch_document:
                response["Content-Disposition"] = "inline; filename=utd.pdf"
            else:
                response["Content-Disposition"] = "attachment; filename=utd.pdf"

        os.remove(temp_excel_path)
        os.remove(temp_pdf_path)
        os.remove(temp_modified_pdf_path)
        os.remove(temp_pdf_path_count)
        os.remove(temp_excel_count)

        return response

        # temp_excel_path = "utd/utils/invoice.xlsx"
        # temp_pdf_path = "utd/utils/invoice.pdf"
        # temp_modified_pdf_path = "utd/utils/invoice_modified.pdf"
        #
        # workbook.save(temp_excel_path)
        #
        # workbook_aspose = Workbook(temp_excel_path)
        # workbook_aspose.save(temp_pdf_path, SaveFormat.PDF)
        #
        # reader = PdfReader(temp_pdf_path)
        # writer = PdfWriter()
        #
        # pages_to_remove = [i for i in range(2, 80)]
        #
        # for i in range(len(reader.pages)):
        #     if i not in pages_to_remove:
        #         writer.add_page(reader.pages[i])
        #
        # with open(temp_modified_pdf_path, "wb") as output_pdf:
        #     writer.write(output_pdf)
        #
        # with open(temp_modified_pdf_path, "rb") as pdf_file:
        #     response = HttpResponse(pdf_file.read(), content_type="application/pdf")
        #     if watch_document:
        #         response["Content-Disposition"] = "inline; filename=invoice.pdf"
        #     else:
        #         response["Content-Disposition"] = "attachment; filename=invoice.pdf"
        #
        # os.remove(temp_excel_path)
        # os.remove(temp_pdf_path)
        # os.remove(temp_modified_pdf_path)
        #
        # return response

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=utd.xlsx"
    workbook.save(response)
    return response
