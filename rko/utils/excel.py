import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.worksheet.table import Table, TableStyleInfo
from aspose.cells import Workbook, HtmlSaveOptions, SaveFormat
from bs4 import BeautifulSoup
import aspose.cells as cells
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime
from PyPDF2 import PdfReader, PdfWriter
import os


def create_rko_excel(data, formset_data, pdf=False, watch_document=False):
    file_path = 'rko/utils/Расходный кассовый ордер №1 от 06.02.2025 г..xlsx'

    workbook = openpyxl.load_workbook(file_path)

    sheet = workbook["Расходный кассовый ордер"]

    for row in sheet.iter_rows():
        if row[0].row == 2:
            sheet.row_dimensions[row[0].row].height = 25

    sheet['A4'] = f'{data["organization"].naming}'
    sheet['BX10'] = f'{data["name"]}'

    date_str = str(data['date'])
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d-%m-%Y")
    sheet['CO10'] = f'{formatted_date}'

    if data["account_debit"]:
        sheet['W15'] = f'{data["account_debit"]}'
    else:
        sheet['W15'] = ''

    if data["account_loan"]:
        sheet['BK15'] = f'{data["account_loan"]}'
    else:
        sheet['BK15'] = ''

    if data["summa"]:
        sheet['BX15'] = f'{data["summa"]}'
    else:
        sheet['BX15'] = ''

    if data["payer"]:
        sheet['I17'] = f'{data["payer"]}'
    else:
        sheet['I17'] = ''

    if data["base"]:
        sheet['L19'] = f'{data["base"]}'
    else:
        sheet['L19'] = ''

    if data["summa"]:
        sheet['H21'] = f'{data["summa"]} рублей'
    else:
        sheet['H21'] = ''

    if data["annex"]:
        sheet['M23'] = f'{data["annex"]}'
    else:
        sheet['M23'] = ''

    if data["organization"].position_at_work:
        sheet['T25'] = f'{data["organization"].position_at_work}'
    else:
        sheet['T25'] = ''

    if data["organization"].supervisor:
        sheet['BZ25'] = f'{data["organization"].supervisor}'
    else:
        sheet['BZ25'] = ''

    if data["organization"].accountant:
        sheet['AW28'] = f'{data["organization"].accountant}'
    else:
        sheet['AW28'] = ''

    if data["passport"]:
        sheet['E37'] = f'{data["passport"]}'
    else:
        sheet['E37'] = ''

    if data["organization"].supervisor:
        sheet['AO42'] = f'{data["organization"].supervisor}'
    else:
        sheet['AO42'] = ''

    if data['organization'].signature and data['is_stamp']:
        image_file = data['organization'].signature

        image_data = image_file.read()

        img_stream = BytesIO(image_data)

        img1 = Image(img_stream)
        img1.width = 70
        img1.height = 25
        sheet.add_image(img1, "AW25")

        img_stream_2 = BytesIO(image_data)

        img2 = Image(img_stream_2)
        img2.width = 70
        img2.height = 25
        sheet.add_image(img2, "P42")

    if pdf:
        temp_excel_path = "rko/utils/invoice.xlsx"
        temp_pdf_path = "rko/utils/invoice.pdf"
        temp_modified_pdf_path = "rko/utils/invoice_modified.pdf"

        workbook.save(temp_excel_path)

        workbook_aspose = Workbook(temp_excel_path)
        workbook_aspose.save(temp_pdf_path, SaveFormat.PDF)

        reader = PdfReader(temp_pdf_path)
        writer = PdfWriter()

        pages_to_remove = [i for i in range(1, 55)]

        for i in range(len(reader.pages)):
            if i not in pages_to_remove:
                writer.add_page(reader.pages[i])

        with open(temp_modified_pdf_path, "wb") as output_pdf:
            writer.write(output_pdf)

        with open(temp_modified_pdf_path, "rb") as pdf_file:
            response = HttpResponse(pdf_file.read(), content_type="application/pdf")
            if watch_document:
                response["Content-Disposition"] = "inline; filename=invoice.pdf"
            else:
                response["Content-Disposition"] = "attachment; filename=invoice.pdf"

        os.remove(temp_excel_path)
        os.remove(temp_pdf_path)
        os.remove(temp_modified_pdf_path)

        return response

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=invoice.xlsx"
    workbook.save(response)
    return response
