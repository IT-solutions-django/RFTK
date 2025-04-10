import convertapi
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.worksheet.table import Table, TableStyleInfo
from aspose.cells import Workbook, HtmlSaveOptions, SaveFormat, License
from bs4 import BeautifulSoup
import aspose.cells as cells
from datetime import datetime
from openpyxl.drawing.image import Image
from io import BytesIO
import locale
from PyPDF2 import PdfReader, PdfWriter
import os
import math

months_russian = [
    'Января', 'Февраля', 'Марта', 'Апреля', 'Мая', 'Июня',
    'Июля', 'Августа', 'Сентября', 'Октября', 'Ноября', 'Декабря'
]


def html_to_excel():
    html_file = "reconciliation/utils/updated_file.html"

    html_load_options = cells.HtmlLoadOptions()

    workbook = cells.Workbook(html_file, html_load_options)

    workbook.save("reconciliation/utils/output.xlsx")


def change_html(count_rows):
    with open("reconciliation/utils/Счет_на_оплату.html", "r",
              encoding="utf-8") as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, "html.parser")

    element_to_remove = soup.find("div",
                                  string=lambda text: text and "Evaluation Only. Created with Aspose.Cells" in text)

    if element_to_remove:
        element_to_remove.decompose()

    target_row = soup.find("td", string=lambda
        text: text and "TestName" in text)

    if target_row:
        parent_row = target_row.find_parent("tr")

        if parent_row:
            num_rows_to_add = count_rows - 1

            if num_rows_to_add >= 1:
                for _ in range(num_rows_to_add):
                    new_row = BeautifulSoup(str(parent_row), "html.parser")

                    parent_row.insert_after(new_row)

            with open("reconciliation/utils/updated_file.html", "w",
                      encoding="utf-8") as file:
                file.write(str(soup))


def excel_to_html():
    file_path = "reconciliation/utils/Акт сверки взаиморасчетов №1 от 27.03.2025 г..xlsx"
    workbook = Workbook(file_path)

    save_options = HtmlSaveOptions()
    save_options.export_active_worksheet_only = True
    save_options.export_images_as_base64 = True

    html_file = "reconciliation/utils/Счет_на_оплату.html"
    workbook.save(html_file, save_options)


def create_reconciliation_excel(data, formset_data, pdf=False, watch_document=False):
    excel_to_html()
    change_html(len(formset_data))
    html_to_excel()

    file_path = 'reconciliation/utils/output.xlsx'

    workbook = openpyxl.load_workbook(file_path)
    workbook.remove(workbook["Evaluation Warning"])

    sheet = workbook["Акт сверки взаиморасчетов"]

    for col in sheet.columns:
        try:
            sheet.column_dimensions[col[0].column_letter].width = 1
        except:
            continue

    for row in sheet.iter_rows():
        if row[0].row in [12]:
            sheet.row_dimensions[row[0].row].height = 22

    start_col = 150
    num_cols = 790

    sheet.delete_cols(start_col, num_cols)

    sheet['A2'] = 'Акт сверки ' + data['name']

    date_str_from = str(data['period_from'])
    date_obj_from = datetime.strptime(date_str_from, "%Y-%m-%d")
    formatted_date_from = date_obj_from.strftime("%d %B %Y").split(' ')

    month_from = date_obj_from.month
    month_russian_from = months_russian[month_from - 1]

    date_str_by = str(data['period_by'])
    date_obj_by = datetime.strptime(date_str_by, "%Y-%m-%d")
    formatted_date_by = date_obj_by.strftime("%d %B %Y").split(' ')

    month_by = date_obj_by.month
    month_russian_by = months_russian[month_by - 1]

    sheet['A4'] = f'взаимных расчетов за период с {formatted_date_from[0]} {month_russian_from} {formatted_date_from[2]} по {formatted_date_by[0]} {month_russian_by} {formatted_date_by[2]}'

    sheet['A6'] = f'между {data["organization"].naming}'

    sheet['A8'] = f'и {data["counterparty"].naming}'

    date_str = str(data['date'])
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d %m %Y").split(' ')
    sheet['A10'] = f'{formatted_date[0]}.{formatted_date[1]}.{formatted_date[2]}'

    sheet['BV10'] = data['place_of_act']

    sheet['A12'] = f'Мы, нижеподписавшиеся, {data["organization"].position_at_work} {data["organization"].naming} {data["organization"].supervisor} с одной стороны, и {data["counterparty"].naming}, с другой стороны, составили настоящий акт сверки в том, что состояние взаимных расчетов по данным учета следующее:'

    sheet['A14'] = f'По данным {data["organization"].naming}, руб.'

    sheet['BD14'] = f'По данным {data["counterparty"].naming}, руб.'

    date_str = str(data['period_from'])
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d %m %Y").split(' ')
    sheet['E16'] = f'Сальдо {formatted_date[0]}.{formatted_date[1]}.{formatted_date[2]}'

    date_str = str(data['period_from'])
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d %m %Y").split(' ')
    sheet['BH16'] = f'Сальдо {formatted_date[0]}.{formatted_date[1]}.{formatted_date[2]}'

    if data["balance_debit"]:
        sheet['AH16'] = f'{data["balance_debit"]}'
    else:
        sheet['AH16'] = "-"

    if data["balance_loan"]:
        sheet['AS16'] = f'{data["balance_loan"]}'
    else:
        sheet['AS16'] = "-"

    if data["balance_loan"]:
        sheet['CJ16'] = f'{data["balance_loan"]}'
    else:
        sheet['CJ16'] = "-"

    if data["balance_debit"]:
        sheet['CU16'] = f'{data["balance_debit"]}'
    else:
        sheet['CU16'] = "-"

    start_table_row = 16
    sum_debit_org = 0
    sum_loan_org = 0
    sum_debit_counterparty = 0
    sum_loan_counterparty = 0
    total_sum_org = 0
    total_sum_counterparty = 0

    max_symbol_line = 15

    height_line = 10

    for idx, table_data in enumerate(formset_data, 1):
        len_name_org = math.ceil(len(table_data['name_operation_org']) / max_symbol_line)
        len_name_cou = math.ceil(len(table_data['name_operation_counterparty']) / max_symbol_line)

        sheet.row_dimensions[start_table_row + idx].height = max(len_name_org, len_name_cou) * height_line

        if table_data['debit_org']:
            sum_debit_org += int(table_data['debit_org'])

        if table_data['loan_org']:
            sum_loan_org += int(table_data['loan_org'])

        if table_data['debit_counterparty']:
            sum_debit_counterparty += int(table_data['debit_counterparty'])

        if table_data['loan_counterparty']:
            sum_loan_counterparty += int(table_data['loan_counterparty'])

        sheet[f'A{start_table_row + idx}'] = f'{idx + 1}'

        sheet[f'BD{start_table_row + idx}'] = f'{idx + 1}'

        sheet[f'E{start_table_row + idx}'] = table_data['name_operation_org']
        sheet[f'AH{start_table_row + idx}'] = table_data['debit_org']
        sheet[f'AS{start_table_row + idx}'] = table_data['loan_org']

        sheet[f'BH{start_table_row + idx}'] = table_data['name_operation_counterparty']
        sheet[f'CJ{start_table_row + idx}'] = table_data['debit_counterparty']
        sheet[f'CU{start_table_row + idx}'] = table_data['loan_counterparty']

    for row in sheet.iter_rows():
        if row[0].row in [start_table_row + len(formset_data) + 5]:
            sheet.row_dimensions[row[0].row].height = 22

    sheet[f'A{start_table_row + len(formset_data) + 1}'] = f'{len(formset_data) + 2}'

    sheet[f'AH{start_table_row + len(formset_data) + 1}'] = f'{sum_debit_org}'

    sheet[f'AS{start_table_row + len(formset_data) + 1}'] = f'{sum_loan_org}'

    sheet[f'BD{start_table_row + len(formset_data) + 1}'] = f'{len(formset_data) + 2}'

    sheet[f'CJ{start_table_row + len(formset_data) + 1}'] = f'{sum_debit_counterparty}'

    sheet[f'CU{start_table_row + len(formset_data) + 1}'] = f'{sum_loan_counterparty}'

    sheet[f'A{start_table_row + len(formset_data) + 2}'] = f'{len(formset_data) + 3}'

    sheet[f'BD{start_table_row + len(formset_data) + 2}'] = f'{len(formset_data) + 3}'

    date_str = str(data['period_by'])
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d %m %Y").split(' ')
    sheet[f'E{start_table_row + len(formset_data) + 2}'] = f'Сальдо на {formatted_date[0]}.{formatted_date[1]}.{formatted_date[2]}'

    if sum_loan_org > sum_debit_org:
        if data['balance_debit']:
            sheet[f'AS{start_table_row + len(formset_data) + 2}'] = f'{sum_loan_org - sum_debit_org + int(data.get("balance_debit", 0))}'
        else:
            sheet[
                f'AS{start_table_row + len(formset_data) + 2}'] = f'{sum_loan_org - sum_debit_org}'
        sheet[
            f'AH{start_table_row + len(formset_data) + 2}'] = ''
        if data['balance_debit']:
            total_sum_org = sum_loan_org - sum_debit_org + int(data.get("balance_debit", 0))
        else:
            total_sum_org = sum_loan_org - sum_debit_org
    elif sum_loan_org <= sum_debit_org:
        if data['balance_debit']:
            sheet[
                f'AH{start_table_row + len(formset_data) + 2}'] = f'{abs(sum_debit_org - sum_loan_org - int(data.get("balance_debit", 0)))}'
        else:
            sheet[
                f'AH{start_table_row + len(formset_data) + 2}'] = f'{abs(sum_debit_org - sum_loan_org)}'
        sheet[
            f'AS{start_table_row + len(formset_data) + 2}'] = ''
        if data['balance_debit']:
            total_sum_org = abs(sum_debit_org - sum_loan_org - int(data.get("balance_debit", 0)))
        else:
            total_sum_org = abs(sum_debit_org - sum_loan_org)

    date_str = str(data['period_by'])
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d %m %Y").split(' ')
    sheet[
        f'BH{start_table_row + len(formset_data) + 2}'] = f'Сальдо на {formatted_date[0]}.{formatted_date[1]}.{formatted_date[2]}'

    if sum_debit_counterparty >= sum_loan_counterparty:
        if data['balance_debit']:
            sheet[
                f'CJ{start_table_row + len(formset_data) + 2}'] = f'{sum_debit_counterparty - sum_loan_counterparty + int(data.get("balance_debit", 0))}'
        else:
            sheet[
                f'CJ{start_table_row + len(formset_data) + 2}'] = f'{sum_debit_counterparty - sum_loan_counterparty}'
        sheet[
            f'CU{start_table_row + len(formset_data) + 2}'] = ''
        if data['balance_debit']:
            total_sum_counterparty = sum_debit_counterparty - sum_loan_counterparty + int(data.get("balance_debit", 0))
        else:
            total_sum_counterparty = sum_debit_counterparty - sum_loan_counterparty
    elif sum_debit_counterparty < sum_loan_counterparty:
        if data['balance_debit']:
            sheet[
                f'CU{start_table_row + len(formset_data) + 2}'] = f'{abs(sum_loan_counterparty - sum_debit_counterparty - int(data.get("balance_debit", 0)))}'
        else:
            sheet[
                f'CU{start_table_row + len(formset_data) + 2}'] = f'{abs(sum_loan_counterparty - sum_debit_counterparty)}'
        sheet[
            f'CJ{start_table_row + len(formset_data) + 2}'] = ''
        if data['balance_debit']:
            total_sum_counterparty = abs(sum_loan_counterparty - sum_debit_counterparty - int(data.get("balance_debit", 0)))
        else:
            total_sum_counterparty = abs(
                sum_loan_counterparty - sum_debit_counterparty)

    sheet[f'A{start_table_row + len(formset_data) + 4}'] = f'По данным {data["organization"].naming}'

    sheet[f'BE{start_table_row + len(formset_data) + 4}'] = f'По данным {data["counterparty"].naming}'

    if sum_loan_org > sum_debit_org:
        sheet[f'A{start_table_row + len(formset_data) + 5}'] = f'на {formatted_date[0]}.{formatted_date[1]}.{formatted_date[2]} задолженность в пользу {data["counterparty"].naming} {total_sum_org} руб.'
    elif sum_loan_org <= sum_debit_org:
        sheet[
            f'A{start_table_row + len(formset_data) + 5}'] = f'на {formatted_date[0]}.{formatted_date[1]}.{formatted_date[2]} задолженность в пользу {data["organization"].naming} {total_sum_org} руб.'

    if sum_debit_counterparty >= sum_loan_counterparty:
        sheet[f'BE{start_table_row + len(formset_data) + 5}'] = f'на {formatted_date[0]}.{formatted_date[1]}.{formatted_date[2]} задолженность в пользу {data["counterparty"].naming} {total_sum_counterparty} руб.'
    elif sum_debit_counterparty < sum_loan_counterparty:
        sheet[
            f'BE{start_table_row + len(formset_data) + 5}'] = f'на {formatted_date[0]}.{formatted_date[1]}.{formatted_date[2]} задолженность в пользу {data["organization"].naming} {total_sum_counterparty} руб.'

    sheet[f'A{start_table_row + len(formset_data) + 7}'] = f'от {data["organization"].naming}'

    sheet[f'BE{start_table_row + len(formset_data) + 7}'] = f'от {data["counterparty"].naming}'

    sheet[f'A{start_table_row + len(formset_data) + 8}'] = f'{data["organization"].position_at_work}'

    sheet[f'BE{start_table_row + len(formset_data) + 8}'] = ''

    sheet[f'R{start_table_row + len(formset_data) + 10}'] = f'{data["organization"].supervisor}'

    sheet[f'BV{start_table_row + len(formset_data) + 10}'] = ''

    # if data['organization'].stamp and data['is_stamp']:
    #     image_file = data['organization'].stamp
    #
    #     image_data = image_file.read()
    #
    #     img_stream1 = BytesIO(image_data)
    #     img1 = Image(img_stream1)
    #     img1.width = 43 * 2.83
    #     img1.height = 43 * 2.83
    #
    #     if len(formset_data) == 1:
    #         sheet.add_image(img1, f"AI{start_table_row + len(formset_data) + 20}")
    #     else:
    #         sheet.add_image(img1, f"AI{start_table_row + len(formset_data) + 25}")
    #
    # if data['organization'].signature and data['is_stamp']:
    #     image_file = data['organization'].signature
    #
    #     image_data = image_file.read()
    #
    #     img_stream = BytesIO(image_data)
    #
    #     img1 = Image(img_stream)
    #     img1.width = 70
    #     img1.height = 40
    #     sheet.add_image(img1, f"AD{start_table_row + len(formset_data) + 14}")
    #
    #     img_stream2 = BytesIO(image_data)
    #     img2 = Image(img_stream2)
    #     img2.width = 70
    #     img2.height = 40
    #     sheet.add_image(img2, f"AD{start_table_row + len(formset_data) + 21}")
    #
    #     img_stream3 = BytesIO(image_data)
    #     img3 = Image(img_stream3)
    #     img3.width = 70
    #     img3.height = 40
    #     sheet.add_image(img3, f"AZ{start_table_row + len(formset_data) + 3}")
    #
    # for row in sheet.iter_rows():
    #     if row[0].row in [start_table_row + len(formset_data) + 3, start_table_row + len(formset_data) + 14,
    #                       start_table_row + len(formset_data) + 21]:
    #         sheet.row_dimensions[row[0].row].height = 20

    if pdf:
        convertapi.api_credentials = 'secret_wBUU4YjxTpfeIPwA'

        temp_excel_path = "reconciliation/utils/invoice.xlsx"
        temp_modified_pdf_path = "reconciliation/utils/invoice_modified.pdf"

        workbook.save(temp_excel_path)

        temp_pdf_path = convertapi.convert('pdf', {
            'File': temp_excel_path,
        }, from_format='xls').save_files('reconciliation/utils')[0]

        reader = PdfReader(temp_pdf_path)
        writer = PdfWriter()

        pages_to_remove = [i for i in range(1, 55)]

        for i in range(len(reader.pages)):
            writer.add_page(reader.pages[i])
            # if i not in pages_to_remove:
            #     writer.add_page(reader.pages[i])

        with open(temp_modified_pdf_path, "wb") as output_pdf:
            writer.write(output_pdf)

        with open(temp_modified_pdf_path, "rb") as pdf_file:
            response = HttpResponse(pdf_file.read(), content_type="application/pdf")
            if watch_document:
                response["Content-Disposition"] = "inline; filename=reconciliation.pdf"
            else:
                response["Content-Disposition"] = "attachment; filename=reconciliation.pdf"

        os.remove(temp_excel_path)
        os.remove(temp_pdf_path)
        os.remove(temp_modified_pdf_path)

        return response

        # temp_excel_path = "sales_receipt/utils/invoice.xlsx"
        # temp_pdf_path = "sales_receipt/utils/invoice.pdf"
        # temp_modified_pdf_path = "sales_receipt/utils/invoice_modified.pdf"
        #
        # workbook.save(temp_excel_path)
        #
        # workbook_aspose = Workbook(temp_excel_path)
        # workbook_aspose.save(temp_pdf_path, SaveFormat.PDF)
        #
        # reader = PdfReader(temp_pdf_path)
        # writer = PdfWriter()
        #
        # pages_to_remove = [i for i in range(1, 55)]
        #
        # for i in range(len(reader.pages)):
        #     if i not in pages_to_remove:
        #         writer.add_page(reader.pages[i])
        #
        # with open(temp_modified_pdf_path, "wb") as output_pdf:
        #     writer.write(output_pdf)
        #
        # with open(temp_modified_pdf_path, "rb") as pdf_file:
        #     response = HttpResponse(pdf_file.read(), content_type="application/pdf")
        #     if watch_document:
        #         response["Content-Disposition"] = "inline; filename=invoice.pdf"
        #     else:
        #         response["Content-Disposition"] = "attachment; filename=invoice.pdf"
        #
        # os.remove(temp_excel_path)
        # os.remove(temp_pdf_path)
        # os.remove(temp_modified_pdf_path)
        #
        # return response

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=reconciliation.xlsx"
    workbook.save(response)
    return response
