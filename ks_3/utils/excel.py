import convertapi
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.worksheet.table import Table, TableStyleInfo
from aspose.cells import Workbook, HtmlSaveOptions, SaveFormat, License
from bs4 import BeautifulSoup
import aspose.cells as cells
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime
from PyPDF2 import PdfReader, PdfWriter
import os


def html_to_excel():
    html_file = "ks_3/utils/updated_file.html"

    html_load_options = cells.HtmlLoadOptions()

    workbook = cells.Workbook(html_file, html_load_options)

    workbook.save("ks_3/utils/output.xlsx")


def change_html(count_rows):
    with open("ks_3/utils/Счет_на_оплату.html", "r",
              encoding="utf-8") as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, "html.parser")

    element_to_remove = soup.find("div",
                                  string=lambda text: text and "Evaluation Only. Created with Aspose.Cells" in text)

    if element_to_remove:
        element_to_remove.decompose()

    target_row = soup.find("td", string=lambda
        text: text and "Лента" in text)

    if target_row:
        parent_row = target_row.find_parent("tr")

        if parent_row:
            num_rows_to_add = count_rows - 1

            if num_rows_to_add >= 1:
                for _ in range(num_rows_to_add):
                    new_row = BeautifulSoup(str(parent_row), "html.parser")

                    parent_row.insert_after(new_row)

            with open("ks_3/utils/updated_file.html",
                      "w",
                      encoding="utf-8") as file:
                file.write(str(soup))


def excel_to_html():
    file_path = "ks_3/utils/КС-3 №1 от 05.02.2025 г..xlsx"
    workbook = Workbook(file_path)

    save_options = HtmlSaveOptions()
    save_options.export_active_worksheet_only = True
    save_options.export_images_as_base64 = True

    html_file = "ks_3/utils/Счет_на_оплату.html"
    workbook.save(html_file, save_options)


def create_ks3_excel(data, formset_data, pdf=False, watch_document=False):
    excel_to_html()
    change_html(len(formset_data))
    html_to_excel()

    file_path = 'ks_3/utils/output.xlsx'

    workbook = openpyxl.load_workbook(file_path)
    workbook.remove(workbook["Evaluation Warning"])

    sheet = workbook["КС-3"]

    for row in sheet.iter_rows():
        if row[0].row in [7, 9]:
            sheet.row_dimensions[row[0].row].height = 22

    start_col = 208
    num_cols = 790

    sheet.delete_cols(start_col, num_cols)

    if data["investor"]:
        sheet['K5'] = f'{data["investor"].naming}, {data["investor"].address}'
    else:
        sheet['K5'] = ''

    sheet['X7'] = f'{data["counterparty"].naming}, {data["counterparty"].address}'
    sheet['Z9'] = f'{data["organization"].naming}, {data["organization"].address}'
    sheet['CQ8'] = ''

    if data["name_construction"]:
        sheet['J11'] = f'{data["name_construction"]}'
    elif data["address_construction"]:
        sheet['J11'] = f'{data["address_construction"]}'
    elif data["name_construction"] and data["address_construction"]:
        sheet['J11'] = f'{data["name_construction"]}, {data["address_construction"]}'
    else:
        sheet['J11'] = ''

    if data["number_agreement"]:
        sheet['CQ14'] = f'{data["number_agreement"]}'
    else:
        sheet['CQ14'] = ''

    if data['date_agreement']:
        date_str = str(data['date_agreement'])
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        formatted_date = date_obj.strftime("%d %m %Y").split(' ')
        sheet['CQ15'] = f'{formatted_date[0]}'
        sheet['CV15'] = f'{formatted_date[1]}'
        sheet['DA15'] = f'{formatted_date[2]}'
    else:
        sheet['CQ15'] = ''
        sheet['CV15'] = ''
        sheet['DA15'] = ''

    sheet['BB20'] = f'{data["name"]}'

    date_str = str(data['date'])
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d-%m-%Y")
    sheet['BQ20'] = f'{formatted_date}'

    if data['period_from']:
        date_str = str(data['period_from'])
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        formatted_date = date_obj.strftime("%d-%m-%Y")
        sheet['CJ20'] = f'{formatted_date}'
    else:
        sheet['CJ20'] = ''

    if data['period_by']:
        date_str = str(data['period_by'])
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        formatted_date = date_obj.strftime("%d-%m-%Y")
        sheet['CU20'] = f'{formatted_date}'
    else:
        sheet['CU20'] = ''

    sheet['BM28'] = ''
    sheet['CB28'] = ''
    sheet['CQ28'] = ''

    start_table_row = 29
    total_sum = 0
    total_nds = 0
    total_sum_with_nds = 0
    total_sum_work = 0
    total_sum_year = 0

    if int(data['nds']) > 0 and int(data['nds']):
        nds = int(data['nds'])
    else:
        nds = 0

    for idx, table_data in enumerate(formset_data, 1):
        sheet[f'A{start_table_row + idx}'] = f'{idx}'
        sheet[f'G{start_table_row + idx}'] = table_data['name']
        sheet[f'BE{start_table_row + idx}'] = table_data['code']
        if table_data['price_from_construction']:
            sheet[f'BM{start_table_row + idx}'] = table_data['price_from_construction']
            total_sum_work += table_data['price_from_construction']
        else:
            sheet[f'BM{start_table_row + idx}'] = '0'
        if table_data['price_from_year']:
            sheet[f'CB{start_table_row + idx}'] = f"{table_data['price_from_year']}"
            total_sum_year += table_data['price_from_year']
        else:
            sheet[f'CB{start_table_row + idx}'] = '0'

        sheet[
            f'CQ{start_table_row + idx}'] = f'{round(float(table_data["price"]) * (float(table_data["quantity"])), 2)}'

        total_sum_with_nds += round(float(table_data["price"]) * (float(table_data["quantity"])), 2)
        total_sum += table_data['amount']
        total_nds += round((float(table_data["price"]) * (float(table_data["quantity"])) * nds * 0.01), 2)

    sheet['BM28'] = f'{total_sum_work}'
    sheet['CB28'] = f'{total_sum_year}'
    sheet['CQ28'] = f'{total_sum}'

    sheet[f'CQ{start_table_row + len(formset_data) + 1}'] = f'{total_sum_with_nds}'
    if data['nds'] == -1 or data['nds'] == '-1':
        sheet[f'CQ{start_table_row + len(formset_data) + 2}'] = 'Без НДС'
    else:
        sheet[f'CQ{start_table_row + len(formset_data) + 2}'] = f'{total_nds}'
    sheet[f'BM{start_table_row + len(formset_data) + 3}'] = 'Всего по акту'
    sheet[f'CQ{start_table_row + len(formset_data) + 3}'] = f'{total_sum}'

    sheet[f'Z{start_table_row + len(formset_data) + 9}'] = f'{data["organization"].position_at_work}'
    sheet[f'BX{start_table_row + len(formset_data) + 9}'] = f'{data["organization"].supervisor}'

    sheet[f'Z{start_table_row + len(formset_data) + 5}'] = ''
    sheet[f'BX{start_table_row + len(formset_data) + 5}'] = ''

    if data['organization'].stamp and data['is_stamp']:
        image_file = data['organization'].stamp
        img = Image(BytesIO(image_file.read()))

        img.width = 45 * 2.83
        img.height = 45 * 2.83

        sheet.add_image(img, f"F{start_table_row + len(formset_data) + 7}")

    if data['organization'].signature and data['is_stamp']:
        image_file = data['organization'].signature

        image_data = image_file.read()

        img_stream = BytesIO(image_data)

        img1 = Image(img_stream)
        img1.width = 80
        img1.height = 30
        sheet.add_image(img1, f"BB{start_table_row + len(formset_data) + 9}")

    if pdf:
        convertapi.api_credentials = 'secret_VEJPjELYZzhUihM6'

        temp_excel_path = "ks_3/utils/invoice.xlsx"
        temp_modified_pdf_path = "ks_3/utils/invoice_modified.pdf"

        workbook.save(temp_excel_path)

        temp_pdf_path = convertapi.convert('pdf', {
            'File': temp_excel_path,
        }, from_format='xls').save_files('ks_3/utils')[0]

        reader = PdfReader(temp_pdf_path)
        writer = PdfWriter()

        pages_to_remove = [i for i in range(1, 55)]

        for i in range(len(reader.pages)):
            writer.add_page(reader.pages[i])
            # if i not in pages_to_remove:
            #     writer.add_page(reader.pages[i])

        with open(temp_modified_pdf_path, "wb") as output_pdf:
            writer.write(output_pdf)

        with open(temp_modified_pdf_path, "rb") as pdf_file:
            response = HttpResponse(pdf_file.read(), content_type="application/pdf")
            if watch_document:
                response["Content-Disposition"] = "inline; filename=КС-3.pdf"
            else:
                response["Content-Disposition"] = "attachment; filename=КС-3.pdf"

        os.remove(temp_excel_path)
        os.remove(temp_pdf_path)
        os.remove(temp_modified_pdf_path)

        return response

        # temp_excel_path = "ks_3/utils/invoice.xlsx"
        # temp_pdf_path = "ks_3/utils/invoice.pdf"
        # temp_modified_pdf_path = "ks_3/utils/invoice_modified.pdf"
        #
        # workbook.save(temp_excel_path)
        #
        # workbook_aspose = Workbook(temp_excel_path)
        # workbook_aspose.save(temp_pdf_path, SaveFormat.PDF)
        #
        # reader = PdfReader(temp_pdf_path)
        # writer = PdfWriter()
        #
        # pages_to_remove = [i for i in range(1, 55)]
        #
        # for i in range(len(reader.pages)):
        #     if i not in pages_to_remove:
        #         writer.add_page(reader.pages[i])
        #
        # with open(temp_modified_pdf_path, "wb") as output_pdf:
        #     writer.write(output_pdf)
        #
        # with open(temp_modified_pdf_path, "rb") as pdf_file:
        #     response = HttpResponse(pdf_file.read(), content_type="application/pdf")
        #     if watch_document:
        #         response["Content-Disposition"] = "inline; filename=invoice.pdf"
        #     else:
        #         response["Content-Disposition"] = "attachment; filename=invoice.pdf"
        #
        # os.remove(temp_excel_path)
        # os.remove(temp_pdf_path)
        # os.remove(temp_modified_pdf_path)
        #
        # return response

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=КС-3.xlsx"
    workbook.save(response)
    return response
