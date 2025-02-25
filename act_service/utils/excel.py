import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.worksheet.table import Table, TableStyleInfo
from aspose.cells import Workbook, HtmlSaveOptions, SaveFormat, License
from bs4 import BeautifulSoup
import aspose.cells as cells
from openpyxl.drawing.image import Image
from io import BytesIO
import logging
from datetime import datetime
from PyPDF2 import PdfReader, PdfWriter
import os
import convertapi

logging.basicConfig(level=logging.DEBUG)


def html_to_excel():
    html_file = "act_service/utils/updated_file.html"

    html_load_options = cells.HtmlLoadOptions()

    workbook = cells.Workbook(html_file, html_load_options)

    workbook.save("act_service/utils/output.xlsx")


def change_html(count_rows):
    with open("act_service/utils/Счет_на_оплату.html", "r",
              encoding="utf-8") as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, "html.parser")

    element_to_remove = soup.find("div",
                                  string=lambda text: text and "Evaluation Only. Created with Aspose.Cells" in text)

    if element_to_remove:
        element_to_remove.decompose()

    target_row = soup.find("td", string=lambda
        text: text and "Beton" in text)

    if target_row:
        parent_row = target_row.find_parent("tr")

        if parent_row:
            num_rows_to_add = count_rows - 1

            if num_rows_to_add >= 1:
                for _ in range(num_rows_to_add):
                    new_row = BeautifulSoup(str(parent_row), "html.parser")

                    parent_row.insert_after(new_row)

            with open("act_service/utils/updated_file.html",
                      "w",
                      encoding="utf-8") as file:
                file.write(str(soup))


def excel_to_html():
    file_path = "act_service/utils/Акт оказания услуг №1 от 05.02.2025 г..xlsx"
    workbook = Workbook(file_path)

    save_options = HtmlSaveOptions()
    save_options.export_active_worksheet_only = True
    save_options.export_images_as_base64 = True

    html_file = "act_service/utils/Счет_на_оплату.html"
    workbook.save(html_file, save_options)


def create_act_service_excel(data, formset_data, pdf=False, watch_document=False):
    try:
        excel_to_html()
        change_html(len(formset_data))
        html_to_excel()

        file_path = 'act_service/utils/output.xlsx'

        workbook = openpyxl.load_workbook(file_path)
        workbook.remove(workbook["Evaluation Warning"])

        sheet = workbook["Акт оказания услуг"]

        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = 1

        for row in sheet.iter_rows():
            if row[0].row == 9:
                sheet.row_dimensions[row[0].row].height = 22

        start_col = 110
        num_cols = 790

        sheet.delete_cols(start_col, num_cols)

        date_str = str(data['date'])
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        formatted_date = date_obj.strftime("%d-%m-%Y")
        sheet['A2'] = f'Акт № {data["name"]} от {formatted_date}'

        if data["agreement"]:
            sheet['A4'] = f'выполненных работ / оказанных услуг по договору {data["agreement"]}'
        else:
            sheet['A4'] = f'выполненных работ / оказанных услуг по договору'

        if data["payment_for"]:
            sheet['A5'] = f'за {data["payment_for"]}'
        else:
            sheet['A5'] = ''

        sheet[
            'Q7'] = f'{data["organization"].naming}, ИНН {data["organization"].inn}, КПП {data["organization"].kpp}, {data["organization"].address}'
        sheet[
            'Q9'] = f'{data["counterparty"].naming}, ИНН {data["counterparty"].inn}, КПП {data["counterparty"].kpp}, {data["counterparty"].address}'

        start_table_row = 11
        total_sum = 0
        total_nds = 0
        total_sum_with_nds = 0
        if int(data['nds']) > 0 and int(data['nds']):
            nds = int(data['nds'])
        else:
            nds = 0

        for idx, table_data in enumerate(formset_data, 1):
            total_sum += table_data['amount']

            sheet[f'A{start_table_row + idx}'] = f'{idx}'
            sheet[f'E{start_table_row + idx}'] = table_data['name']
            sheet[f'BB{start_table_row + idx}'] = table_data['quantity']
            sheet[f'BM{start_table_row + idx}'] = table_data['unit_of_measurement']
            sheet[f'BW{start_table_row + idx}'] = f'{table_data["price"]}'
            sheet[
                f'CN{start_table_row + idx}'] = f'{round(float(table_data["price"]) * float(table_data["quantity"]), 2)}'
            total_sum_with_nds += round(float(table_data["price"]) * float(table_data["quantity"]), 2)
            total_nds += round(float(table_data["price"]) * float(table_data['quantity'] * nds * 0.01), 2)

        sheet[f'CN{start_table_row + len(formset_data) + 1}'] = f'{total_sum_with_nds}'
        if nds > 0:
            sheet[f'A{start_table_row + len(formset_data) + 2}'] = f'Сумма НДС {data["nds"]}%: {total_nds} руб.'
        else:
            sheet[f'A{start_table_row + len(formset_data) + 2}'] = 'Без НДС'

        sheet[
            f'A{start_table_row + len(formset_data) + 4}'] = f'Всего оказано услуг {len(formset_data)} на сумму {total_sum} руб.'
        sheet[
            f'A{start_table_row + len(formset_data) + 6}'] = ''

        sheet[f'A{start_table_row + len(formset_data) + 11}'] = f'{data["organization"].position_at_work}'
        sheet[f'Q{start_table_row + len(formset_data) + 14}'] = f'{data["organization"].supervisor}'

        sheet[f'BD{start_table_row + len(formset_data) + 11}'] = ''
        sheet[f'BT{start_table_row + len(formset_data) + 14}'] = ''

        if data['organization'].stamp and data['is_stamp']:
            image_file = data['organization'].stamp
            img = Image(BytesIO(image_file.read()))

            img.width = 45 * 2.83
            img.height = 45 * 2.83

            sheet.add_image(img, f"N{start_table_row + len(formset_data) + 12}")

        if data['organization'].signature and data['is_stamp']:
            image_file = data['organization'].signature

            image_data = image_file.read()

            img_stream = BytesIO(image_data)

            img1 = Image(img_stream)
            img1.width = 70
            img1.height = 25
            sheet.add_image(img1, f"A{start_table_row + len(formset_data) + 14}")

        if pdf:
            convertapi.api_credentials = 'secret_VEJPjELYZzhUihM6'

            temp_excel_path = "act_service/utils/invoice.xlsx"
            temp_modified_pdf_path = "act_service/utils/invoice_modified.pdf"

            workbook.save(temp_excel_path)

            temp_pdf_path = convertapi.convert('pdf', {
                'File': temp_excel_path,
            }, from_format='xls').save_files('act_service/utils')[0]

            reader = PdfReader(temp_pdf_path)
            writer = PdfWriter()

            pages_to_remove = [i for i in range(1, 55)]

            for i in range(len(reader.pages)):
                writer.add_page(reader.pages[i])
                # if i not in pages_to_remove:
                #     writer.add_page(reader.pages[i])

            with open(temp_modified_pdf_path, "wb") as output_pdf:
                writer.write(output_pdf)

            with open(temp_modified_pdf_path, "rb") as pdf_file:
                response = HttpResponse(pdf_file.read(), content_type="application/pdf")
                if watch_document:
                    response["Content-Disposition"] = "inline; filename=invoice.pdf"
                else:
                    response["Content-Disposition"] = "attachment; filename=invoice.pdf"

            os.remove(temp_excel_path)
            os.remove(temp_pdf_path)
            os.remove(temp_modified_pdf_path)

            return response

            # temp_excel_path = "act_service/utils/invoice.xlsx"
            # temp_pdf_path = "act_service/utils/invoice.pdf"
            # temp_modified_pdf_path = "act_service/utils/invoice_modified.pdf"
            #
            # workbook.save(temp_excel_path)
            #
            # workbook_aspose = Workbook(temp_excel_path)
            # workbook_aspose.save(temp_pdf_path, SaveFormat.PDF)
            #
            # reader = PdfReader(temp_pdf_path)
            # writer = PdfWriter()
            #
            # pages_to_remove = [i for i in range(1, 55)]
            #
            # for i in range(len(reader.pages)):
            #     if i not in pages_to_remove:
            #         writer.add_page(reader.pages[i])
            #
            # with open(temp_modified_pdf_path, "wb") as output_pdf:
            #     writer.write(output_pdf)
            #
            # with open(temp_modified_pdf_path, "rb") as pdf_file:
            #     response = HttpResponse(pdf_file.read(), content_type="application/pdf")
            #     if watch_document:
            #         response["Content-Disposition"] = "inline; filename=invoice.pdf"
            #     else:
            #         response["Content-Disposition"] = "attachment; filename=invoice.pdf"
            #
            # os.remove(temp_excel_path)
            # os.remove(temp_pdf_path)
            # os.remove(temp_modified_pdf_path)
            #
            # return response

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename=invoice.xlsx"
        workbook.save(response)
        return response
    except Exception as e:
        logging.error(f"Ошибка в create_invoice_excel: {e}")
