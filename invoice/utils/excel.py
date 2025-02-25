import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.worksheet.table import Table, TableStyleInfo
from aspose.cells import Workbook, HtmlSaveOptions, SaveFormat, PdfSaveOptions, PaperSizeType, License
from bs4 import BeautifulSoup
import aspose.cells as cells
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime
from PyPDF2 import PdfReader, PdfWriter
import os

# license_as = License()
# license_as.set_license("lic/Aspose.TotalforPythonvia.NET.lic")


def html_to_excel():
    html_file = "invoice/utils/updated_file.html"

    html_load_options = cells.HtmlLoadOptions()

    workbook = cells.Workbook(html_file, html_load_options)

    workbook.save("invoice/utils/output.xlsx")


def change_html(count_rows):
    with open("invoice/utils/Счет_на_оплату.html", "r",
              encoding="utf-8") as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, "html.parser")

    element_to_remove = soup.find("div",
                                  string=lambda text: text and "Evaluation Only. Created with Aspose.Cells" in text)

    if element_to_remove:
        element_to_remove.decompose()

    target_row = soup.find("td", string=lambda
        text: text and "Лента-герметик самоклеящаяся Технониколь Никобенд, красная, 15х1000 см" in text)

    if target_row:
        parent_row = target_row.find_parent("tr")

        if parent_row:
            num_rows_to_add = count_rows - 1

            if num_rows_to_add >= 1:
                for _ in range(num_rows_to_add):
                    new_row = BeautifulSoup(str(parent_row), "html.parser")

                    parent_row.insert_after(new_row)

            with open("invoice/utils/updated_file.html", "w",
                      encoding="utf-8") as file:
                file.write(str(soup))


def excel_to_html():
    file_path = "invoice/utils/Счет на оплату №3 от 31.01.2025 г..xlsx"
    workbook = Workbook(file_path)

    save_options = HtmlSaveOptions()
    save_options.export_active_worksheet_only = True
    save_options.export_images_as_base64 = True

    html_file = "invoice/utils/Счет_на_оплату.html"
    workbook.save(html_file, save_options)


def create_invoice_excel(data, organization_data, formset_data, pdf=False, watch_document=False):
    excel_to_html()
    change_html(len(formset_data))
    html_to_excel()

    file_path = 'invoice/utils/output.xlsx'

    workbook = openpyxl.load_workbook(file_path)
    workbook.remove(workbook["Evaluation Warning"])

    sheet = workbook["Счет на оплату"]

    for row in sheet.iter_rows():
        if row[0].row == 11:
            sheet.row_dimensions[row[0].row].height = 85

    sheet['A1'] = organization_data['name']
    sheet['A2'] = organization_data['address']
    sheet['A3'] = f"ИНН/КПП {organization_data['inn']} / {organization_data['kpp']}"
    sheet['A4'] = f"ОГРН {organization_data['ogrn']}"

    sheet['A8'] = organization_data['name']

    sheet['A10'] = organization_data['name']
    if data['bank_organization']:
        sheet[
            'A11'] = f"{organization_data['address']}\nИНН/КПП {organization_data['kpp']}/{organization_data['kpp']}\nОГРН {organization_data['ogrn']}\n{data['bank_organization'].naming}, {data['bank_organization'].location}\nКор. счет {data['bank_organization'].correspondent_account}\nБИК {data['bank_organization'].bic}"
    else:
        sheet[
            'A11'] = f"{organization_data['address']}\nИНН/КПП {organization_data['kpp']}/{organization_data['kpp']}\nОГРН {organization_data['ogrn']}"

    sheet['AY10'] = f"{data['counterparty'].naming}"
    if data['bank_counterparty']:
        sheet[
            'AY11'] = f"{data['counterparty'].address}\nИНН {data['counterparty'].inn}\nОГРН {data['counterparty'].ogrn}\n{data['bank_counterparty'].naming}, {data['bank_counterparty'].location}\nКор. счет {data['bank_counterparty'].correspondent_account}\nБИК {data['bank_counterparty'].bic}"
    else:
        sheet[
            'AY11'] = f"{data['counterparty'].address}\nИНН {data['counterparty'].inn}\nОГРН {data['counterparty'].ogrn}"

    date_str = str(data['date'])
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d-%m-%Y")
    sheet['A13'] = f"Счет №{data['name']} от {formatted_date}"

    if data['payment_for']:
        sheet['A15'] = f"за {data['payment_for']}"
    else:
        sheet['A15'] = ''

    start_table_row = 17

    sheet[f'CG{start_table_row}'] = 'Скидка'

    total_sum = 0

    if int(data['nds']) > 0 and int(data['nds']):
        nds = int(data['nds'])
    else:
        nds = 0

    for idx, table_data in enumerate(formset_data, 1):
        total_sum += table_data["amount"]

        sheet[f'A{start_table_row + idx}'] = f'{idx}'
        sheet[f'E{start_table_row + idx}'] = f'{table_data["name"]}'
        sheet[f'BB{start_table_row + idx}'] = f'{table_data["price"]}'
        sheet[f'BP{start_table_row + idx}'] = f'{table_data["quantity"]}'
        sheet[f'BY{start_table_row + idx}'] = f'{table_data["unit_of_measurement"]}'
        if table_data["discount"]:
            sheet[f'CG{start_table_row + idx}'] = f'{table_data["discount"]}'
        else:
            sheet[f'CG{start_table_row + idx}'] = ''

        sheet[f'CN{start_table_row + idx}'] = f'{table_data["amount"]}'

    sheet[f'A{start_table_row + len(formset_data) + 1}'] = 'Итого'
    sheet[f'CN{start_table_row + len(formset_data) + 1}'] = f'{total_sum}'
    if data["agreement"]:
        sheet[f'A{start_table_row + len(formset_data) + 2}'] = f'{data["agreement"]}'
    else:
        sheet[f'A{start_table_row + len(formset_data) + 2}'] = ''

    if data['purpose_of_payment']:
        sheet[f'A{start_table_row + len(formset_data) + 3}'] = f'{data["purpose_of_payment"]}'

    sheet[f'A{start_table_row + len(formset_data) + 7}'] = organization_data['position_at_work']
    sheet[f'AI{start_table_row + len(formset_data) + 7}'] = organization_data['supervisor']
    sheet[f'AI{start_table_row + len(formset_data) + 9}'] = organization_data['accountant']

    if data['organization'].stamp and data['is_stamp']:
        image_file = data['organization'].stamp
        img = Image(BytesIO(image_file.read()))

        img.width = 45 * 2.83
        img.height = 45 * 2.83

        sheet.add_image(img, f"BO{start_table_row + len(formset_data) + 8}")

    if data['organization'].signature and data['is_stamp']:
        image_file = data['organization'].signature

        image_data = image_file.read()

        img_stream = BytesIO(image_data)

        img1 = Image(img_stream)
        img1.width = 70
        img1.height = 35
        sheet.add_image(img1, f"BY{start_table_row + len(formset_data) + 7}")

    if pdf:
        temp_excel_path = "invoice/utils/invoice.xlsx"
        temp_pdf_path = "invoice/utils/invoice.pdf"
        temp_modified_pdf_path = "invoice/utils/invoice_modified.pdf"

        workbook.save(temp_excel_path)

        workbook_aspose = Workbook(temp_excel_path)
        workbook_aspose.save(temp_pdf_path, SaveFormat.PDF)

        reader = PdfReader(temp_pdf_path)
        writer = PdfWriter()

        pages_to_remove = [i for i in range(1, 55)]

        for i in range(len(reader.pages)):
            if i not in pages_to_remove:
                writer.add_page(reader.pages[i])

        with open(temp_modified_pdf_path, "wb") as output_pdf:
            writer.write(output_pdf)

        with open(temp_modified_pdf_path, "rb") as pdf_file:
            response = HttpResponse(pdf_file.read(), content_type="application/pdf")
            if watch_document:
                response["Content-Disposition"] = "inline; filename=invoice.pdf"
            else:
                response["Content-Disposition"] = "attachment; filename=invoice.pdf"

        os.remove(temp_excel_path)
        os.remove(temp_pdf_path)
        os.remove(temp_modified_pdf_path)

        return response

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=invoice.xlsx"
    workbook.save(response)
    return response
