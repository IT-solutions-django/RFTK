import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.worksheet.table import Table, TableStyleInfo
from aspose.cells import Workbook, HtmlSaveOptions
from bs4 import BeautifulSoup
import aspose.cells as cells
from datetime import datetime
from openpyxl.drawing.image import Image
from io import BytesIO


def html_to_excel():
    html_file = "packing_list/utils/updated_file.html"

    html_load_options = cells.HtmlLoadOptions()

    workbook = cells.Workbook(html_file, html_load_options)

    workbook.save("packing_list/utils/output.xlsx")


def change_html(count_rows):
    with open("packing_list/utils/Счет_на_оплату.html", "r",
              encoding="utf-8") as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, "html.parser")

    element_to_remove = soup.find("div",
                                  string=lambda text: text and "Evaluation Only. Created with Aspose.Cells" in text)

    if element_to_remove:
        element_to_remove.decompose()

    target_row = soup.find("td", string=lambda
        text: text and "Чипсы" in text)

    if target_row:
        parent_row = target_row.find_parent("tr")

        if parent_row:
            num_rows_to_add = count_rows - 1

            if num_rows_to_add >= 1:
                for _ in range(num_rows_to_add):
                    new_row = BeautifulSoup(str(parent_row), "html.parser")

                    parent_row.insert_after(new_row)

            with open("/packing_list/utils/updated_file.html", "w",
                      encoding="utf-8") as file:
                file.write(str(soup))


def excel_to_html():
    file_path = "packing_list/utils/Товарная накладная №1 от 01.02.2025 г.-2.xlsx"
    workbook = Workbook(file_path)

    save_options = HtmlSaveOptions()
    save_options.export_active_worksheet_only = True
    save_options.export_images_as_base64 = True

    html_file = "packing_list/utils/Счет_на_оплату.html"
    workbook.save(html_file, save_options)


def create_packing_list_excel(data, formset_data):
    excel_to_html()
    change_html(len(formset_data))
    html_to_excel()

    file_path = 'packing_list/utils/output.xlsx'

    workbook = openpyxl.load_workbook(file_path)

    sheet = workbook["Товарная накладная"]

    sheet['A4'] = f'{data['shipper'].naming}, ИНН: {data['shipper'].inn}, {data['shipper'].address}'
    sheet['EW4'] = ''
    sheet['A6'] = data['structural_division']
    sheet['Q9'] = f'{data['consignee'].naming}, ИНН: {data['consignee'].inn}, {data['consignee'].address}'
    sheet[
        'Q10'] = f'{data['organization'].naming}, ИНН: {data['organization'].inn}, КПП: {data['organization'].kpp}, {data['organization'].address}, {data['bank_organization'].naming}, {data['bank_organization'].location}, БИК {data['bank_organization'].bic}, к/с {data['bank_organization'].correspondent_account}, р/с {data['bank_organization'].current_account}'
    sheet['EW10'] = ''
    sheet[
        'Q11'] = f'{data['counterparty'].naming}, ИНН {data['counterparty'].inn}, {data['organization'].address}, {data['bank_counterparty'].naming}, {data['bank_counterparty'].location}, БИК {data['bank_counterparty'].bic}, к/с {data['bank_counterparty'].correspondent_account}'
    sheet['Q12'] = data['base']
    sheet['EW12'] = data['number_base']
    sheet['EW13'] = data['date_base']
    sheet['EW14'] = data['packing_list']
    sheet['EW15'] = data['date_packing_list']
    sheet['BU16'] = data['name']
    sheet['CL16'] = data['date']

    start_table_row = 23
    total_sum = 0
    total_quantity = 0
    total_gw = 0
    total_nw = 0

    for idx, table_data in enumerate(formset_data, 1):
        total_sum += table_data['amount']
        total_quantity += table_data['quantity']
        total_gw += table_data['gross_weight']
        total_nw += table_data['net_weight']

        sheet[f'A{start_table_row + idx}'] = f'{idx}'
        sheet[f'I{start_table_row + idx}'] = table_data['name']
        sheet[f'BB{start_table_row + idx}'] = table_data['product_code']
        sheet[f'BG{start_table_row + idx}'] = table_data['unit_of_measurement']
        sheet[f'BO{start_table_row + idx}'] = table_data['unit_of_measurement']
        sheet[f'BU{start_table_row + idx}'] = table_data['type_of_packaging']
        sheet[f'CD{start_table_row + idx}'] = table_data['quantity']
        sheet[f'CK{start_table_row + idx}'] = f'{table_data['quantity']}'
        sheet[f'CQ{start_table_row + idx}'] = f'{table_data['gross_weight']}'
        sheet[f'CX{start_table_row + idx}'] = f'{table_data['net_weight']}'
        sheet[f'DH{start_table_row + idx}'] = table_data['price']
        sheet[f'DR{start_table_row + idx}'] = table_data['amount']
        sheet[f'EW{start_table_row + idx}'] = table_data['amount']

    sheet[f'CK{start_table_row + len(formset_data) + 1}'] = f'{total_quantity}'
    sheet[f'CQ{start_table_row + len(formset_data) + 1}'] = f'{total_gw}'
    sheet[f'CX{start_table_row + len(formset_data) + 1}'] = f'{total_nw}'
    sheet[f'DR{start_table_row + len(formset_data) + 1}'] = f'{total_sum}'
    sheet[f'EW{start_table_row + len(formset_data) + 1}'] = f'{total_sum}'

    sheet[f'U{start_table_row + len(formset_data) + 4}'] = f'{len(formset_data)}'
    sheet[f'U{start_table_row + len(formset_data) + 8}'] = f'{total_quantity}'
    sheet[f'CM{start_table_row + len(formset_data) + 8}'] = f'{total_gw} тысяч кг'
    sheet[f'ES{start_table_row + len(formset_data) + 7}'] = f'{total_gw}'

    sheet[f'Y{start_table_row + len(formset_data) + 12}'] = f'{total_sum}'

    sheet[f'S{start_table_row + len(formset_data) + 14}'] = data['organization'].position_at_work
    sheet[f'BC{start_table_row + len(formset_data) + 14}'] = data['organization'].supervisor

    sheet[f'BC{start_table_row + len(formset_data) + 16}'] = data['organization'].accountant

    sheet[f'S{start_table_row + len(formset_data) + 18}'] = data['organization'].position_at_work
    sheet[f'BC{start_table_row + len(formset_data) + 18}'] = data['organization'].supervisor

    date_str = str(data['shipping_date'])
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d %B %Y").split(' ')
    sheet[f'AH{start_table_row + len(formset_data) + 20}'] = formatted_date[0]
    sheet[f'AM{start_table_row + len(formset_data) + 20}'] = formatted_date[1]
    sheet[f'BG{start_table_row + len(formset_data) + 20}'] = formatted_date[2]

    date_str = str(data['date_of_receipt'])
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d %B %Y").split(' ')
    sheet[f'DM{start_table_row + len(formset_data) + 20}'] = formatted_date[0]
    sheet[f'DR{start_table_row + len(formset_data) + 20}'] = formatted_date[1]
    sheet[f'EL{start_table_row + len(formset_data) + 20}'] = formatted_date[2]

    if data['organization'].stamp:
        image_file = data['organization'].stamp

        image_data = image_file.read()

        img_stream1 = BytesIO(image_data)
        img1 = Image(img_stream1)
        img1.width = 50
        img1.height = 50

        sheet.add_image(img1, f"S{start_table_row + len(formset_data) + 20}")

        img_stream2 = BytesIO(image_data)
        img2 = Image(img_stream2)
        img2.width = 50
        img2.height = 50

        sheet.add_image(img2, f"CX{start_table_row + len(formset_data) + 20}")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=invoice.xlsx"
    workbook.save(response)
    return response
