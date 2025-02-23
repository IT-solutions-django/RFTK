import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.worksheet.table import Table, TableStyleInfo
from aspose.cells import Workbook, HtmlSaveOptions, SaveFormat
from bs4 import BeautifulSoup
import aspose.cells as cells
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime
from PyPDF2 import PdfReader, PdfWriter
import os


def html_to_excel():
    html_file = "outlay/utils/updated_file.html"

    html_load_options = cells.HtmlLoadOptions()

    workbook = cells.Workbook(html_file, html_load_options)

    workbook.save("outlay/utils/output.xlsx")


def change_html(count_rows):
    with open("outlay/utils/Счет_на_оплату.html", "r",
              encoding="utf-8") as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, "html.parser")

    element_to_remove = soup.find("div",
                                  string=lambda text: text and "Evaluation Only. Created with Aspose.Cells" in text)

    if element_to_remove:
        element_to_remove.decompose()

    target_row = soup.find("td", string=lambda
        text: text and "Блок газобетонный" in text)

    if target_row:
        parent_row = target_row.find_parent("tr")

        if parent_row:
            num_rows_to_add = count_rows - 1

            if num_rows_to_add >= 1:
                for _ in range(num_rows_to_add):
                    new_row = BeautifulSoup(str(parent_row), "html.parser")

                    parent_row.insert_after(new_row)

            with open("outlay/utils/updated_file.html",
                      "w",
                      encoding="utf-8") as file:
                file.write(str(soup))


def excel_to_html():
    file_path = "outlay/utils/Сметная стоимость №1 от 03.02.2025.xlsx"
    workbook = Workbook(file_path)

    save_options = HtmlSaveOptions()
    save_options.export_active_worksheet_only = True
    save_options.export_images_as_base64 = True

    html_file = "outlay/utils/Счет_на_оплату.html"
    workbook.save(html_file, save_options)


def create_outlay_excel(data, formset_data, pdf=False, watch_document=False):
    excel_to_html()
    change_html(len(formset_data))
    html_to_excel()

    file_path = 'outlay/utils/output.xlsx'

    workbook = openpyxl.load_workbook(file_path)
    workbook.remove(workbook["Evaluation Warning"])

    sheet = workbook["Смета"]

    sheet['BT1'] = f'Приложение № {data["number_outlay"]}'

    date_str = str(data['date'])
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d-%m-%Y")
    sheet['BT2'] = f'от {formatted_date}'

    sheet['BT3'] = data['base']
    sheet['A5'] = data['name']
    sheet['A7'] = data['name_construction']
    if data["address"]:
        sheet['A8'] = f'по адресу {data["address"]}'
    else:
        sheet['A8'] = ''

    if data["work_time"]:
        sheet['A10'] = f'Срок выполнения работ: {data["work_time"]}'
    else:
        sheet['A10'] = ''

    start_table_row = 12
    total_sum = 0
    total_nds = 0
    if int(data['nds']) > 0 and int(data['nds']):
        nds = int(data['nds'])
    else:
        nds = 0

    for idx, table_data in enumerate(formset_data, 1):
        sheet[f'A{start_table_row + idx}'] = f'{idx}'
        sheet[f'E{start_table_row + idx}'] = table_data['name']
        sheet[f'BL{start_table_row + idx}'] = table_data['unit_of_measurement']
        sheet[f'BS{start_table_row + idx}'] = table_data['quantity']
        sheet[f'CB{start_table_row + idx}'] = table_data['price']

        sheet[f'CQ{start_table_row + idx}'] = f'{round(float(table_data["quantity"]) * float(table_data["price"]), 2)}'
        total_sum += table_data['amount']
        total_nds += round((float(table_data["quantity"]) * float(table_data["price"]) * nds * 0.01), 2)

    sheet[f'CQ{start_table_row + len(formset_data) + 1}'] = f'{total_sum}'
    if data["nds"] == -1 or data["nds"] == '-1':
        sheet[f'A{start_table_row + len(formset_data) + 2}'] = f'Без НДС'
        sheet[f'CQ{start_table_row + len(formset_data) + 2}'] = ''
    else:
        sheet[f'A{start_table_row + len(formset_data) + 2}'] = f'НДС {data["nds"]}%:'
        sheet[f'CQ{start_table_row + len(formset_data) + 2}'] = f'{total_nds}'
    sheet[f'A{start_table_row + len(formset_data) + 3}'] = ''
    sheet[f'CQ{start_table_row + len(formset_data) + 3}'] = ''
    sheet[f'A{start_table_row + len(formset_data) + 5}'] = f'Итого: {total_sum} рублей'

    sheet[
        f'BE{start_table_row + len(formset_data) + 9}'] = f'{data["organization"].naming}, ИНН/КПП {data["organization"].inn} / {data["organization"].kpp}'
    sheet[f'BE{start_table_row + len(formset_data) + 10}'] = f'{data["organization"].position_at_work}'
    sheet[f'BS{start_table_row + len(formset_data) + 11}'] = f'{data["organization"].supervisor}'

    sheet[
        f'B{start_table_row + len(formset_data) + 9}'] = f'{data["counterparty"].naming}, ИНН/КПП {data["counterparty"].inn} / {data["counterparty"].kpp}'
    sheet[f'B{start_table_row + len(formset_data) + 10}'] = ''
    sheet[f'Q{start_table_row + len(formset_data) + 11}'] = ''

    if data['organization'].stamp:
        image_file = data['organization'].stamp
        img = Image(BytesIO(image_file.read()))

        img.width = 45 * 2.83
        img.height = 45 * 2.83

        sheet.add_image(img, f"CH{start_table_row + len(formset_data) + 11}")

    if data['organization'].signature and data['is_stamp']:
        image_file = data['organization'].signature

        image_data = image_file.read()

        img_stream = BytesIO(image_data)

        img1 = Image(img_stream)
        img1.width = 70
        img1.height = 25
        sheet.add_image(img1, f"BE{start_table_row + len(formset_data) + 11}")

    if pdf:
        temp_excel_path = "outlay/utils/invoice.xlsx"
        temp_pdf_path = "outlay/utils/invoice.pdf"
        temp_modified_pdf_path = "outlay/utils/invoice_modified.pdf"

        workbook.save(temp_excel_path)

        workbook_aspose = Workbook(temp_excel_path)
        workbook_aspose.save(temp_pdf_path, SaveFormat.PDF)

        reader = PdfReader(temp_pdf_path)
        writer = PdfWriter()

        pages_to_remove = [i for i in range(1, 55)]

        for i in range(len(reader.pages)):
            if i not in pages_to_remove:
                writer.add_page(reader.pages[i])

        with open(temp_modified_pdf_path, "wb") as output_pdf:
            writer.write(output_pdf)

        with open(temp_modified_pdf_path, "rb") as pdf_file:
            response = HttpResponse(pdf_file.read(), content_type="application/pdf")
            if watch_document:
                response["Content-Disposition"] = "inline; filename=invoice.pdf"
            else:
                response["Content-Disposition"] = "attachment; filename=invoice.pdf"

        os.remove(temp_excel_path)
        os.remove(temp_pdf_path)
        os.remove(temp_modified_pdf_path)

        return response

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=invoice.xlsx"
    workbook.save(response)
    return response
