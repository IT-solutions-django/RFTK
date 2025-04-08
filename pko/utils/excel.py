import convertapi
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.worksheet.table import Table, TableStyleInfo
from aspose.cells import Workbook, HtmlSaveOptions, SaveFormat, License
from bs4 import BeautifulSoup
import aspose.cells as cells
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime
from PyPDF2 import PdfReader, PdfWriter
import os


def create_pko_excel(data, formset_data, pdf=False, watch_document=False):
    file_path = 'pko/utils/Приходный кассовый ордер №1 от 06.02.2025 г..xlsx'

    workbook = openpyxl.load_workbook(file_path)

    sheet = workbook["Приходный кассовый ордер"]

    for row in sheet.iter_rows():
        if row[0].row == 2:
            sheet.row_dimensions[row[0].row].height = 25

    start_col = 208
    num_cols = 790

    sheet.delete_cols(start_col, num_cols)

    sheet['A4'] = f'{data["organization"].naming}'
    sheet['AN9'] = f'{data["name"]}'

    date_str = str(data['date'])
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d-%m-%Y")
    sheet['BA9'] = f'{formatted_date}'

    if data["account_debit"]:
        sheet['A15'] = f'{data["account_debit"]}'
    else:
        sheet['A15'] = ''

    if data["account_loan"]:
        sheet['V15'] = f'{data["account_loan"]}'
    else:
        sheet['V15'] = ''

    if data["summa"]:
        sheet['AP15'] = f'{data["summa"]}'
    else:
        sheet['AP15'] = ''

    if data["payer"]:
        sheet['L17'] = f'{data["payer"]}'
    else:
        sheet['L17'] = ''

    if data["base"]:
        sheet['L19'] = f'{data["base"]}'
    else:
        sheet['L19'] = ''

    if data["summa"]:
        sheet['I23'] = f'{data["summa"]} рублей'
    else:
        sheet['I23'] = ''

    if int(data['nds']) and int(data['nds']) > 0 and data["summa"]:
        nds = round(float(data["summa"]) * int(data['nds']) * 0.01, 2)
        sheet['A25'] = f'В том числе НДС({data["nds"]}%): {nds}'

    if data["annex"]:
        sheet['M27'] = f'{data["annex"]}'
    else:
        sheet['M27'] = ''

    if data["organization"].accountant:
        sheet['AK32'] = f'{data["organization"].accountant}'
    else:
        sheet['AK32'] = ''

    if data["organization"].supervisor:
        sheet['AK35'] = f'{data["organization"].supervisor}'
    else:
        sheet['AK35'] = ''

    sheet['BR2'] = f'{data["organization"].naming}'

    date_str = str(data['date'])
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d-%m-%Y")
    sheet['BR6'] = f'к приходному кассовому ордеру № {data["name"]} от {formatted_date}'

    if data["payer"]:
        sheet['CC8'] = f'{data["payer"]}'
    else:
        sheet['CC8'] = ''

    if data["base"]:
        sheet['CC11'] = f'{data["base"]}'
    else:
        sheet['CC11'] = ''

    if data["summa"]:
        sheet['CC13'] = f'{data["summa"]} рублей'
    else:
        sheet['CC13'] = ''

    if int(data['nds']) and int(data['nds']) > 0 and data["summa"]:
        nds = round(float(data["summa"]) * int(data['nds']) * 0.01, 2)
        sheet['BR14'] = f'В том числе НДС({data["nds"]}%): {nds}'

    if data["organization"].accountant:
        sheet['BR29'] = f'{data["organization"].accountant}'
    else:
        sheet['BR29'] = ''

    if data["organization"].supervisor:
        sheet['BR35'] = f'{data["organization"].supervisor}'
    else:
        sheet['BR35'] = ''

    if data['organization'].stamp and data['is_stamp']:
        image_file = data['organization'].stamp
        img = Image(BytesIO(image_file.read()))

        img.width = 45 * 2.83
        img.height = 45 * 2.83

        sheet.add_image(img, "BR17")

    if data['organization'].signature and data['is_stamp']:
        image_file = data['organization'].signature

        image_data = image_file.read()

        img_stream = BytesIO(image_data)

        img1 = Image(img_stream)
        img1.width = 70
        img1.height = 25
        sheet.add_image(img1, "T35")

        img_stream_2 = BytesIO(image_data)

        img2 = Image(img_stream_2)
        img2.width = 70
        img2.height = 25
        sheet.add_image(img2, "BR33")

    if pdf:
        convertapi.api_credentials = 'secret_wBUU4YjxTpfeIPwA'

        temp_excel_path = "pko/utils/invoice.xlsx"
        temp_modified_pdf_path = "pko/utils/invoice_modified.pdf"

        workbook.save(temp_excel_path)

        temp_pdf_path = convertapi.convert('pdf', {
            'File': temp_excel_path,
        }, from_format='xls').save_files('pko/utils')[0]

        reader = PdfReader(temp_pdf_path)
        writer = PdfWriter()

        pages_to_remove = [i for i in range(1, 55)]

        for i in range(len(reader.pages)):
            writer.add_page(reader.pages[i])
            # if i not in pages_to_remove:
            #     writer.add_page(reader.pages[i])

        with open(temp_modified_pdf_path, "wb") as output_pdf:
            writer.write(output_pdf)

        with open(temp_modified_pdf_path, "rb") as pdf_file:
            response = HttpResponse(pdf_file.read(), content_type="application/pdf")
            if watch_document:
                response["Content-Disposition"] = "inline; filename=pko.pdf"
            else:
                response["Content-Disposition"] = "attachment; filename=pko.pdf"

        os.remove(temp_excel_path)
        os.remove(temp_pdf_path)
        os.remove(temp_modified_pdf_path)

        return response

        # temp_excel_path = "pko/utils/invoice.xlsx"
        # temp_pdf_path = "pko/utils/invoice.pdf"
        # temp_modified_pdf_path = "pko/utils/invoice_modified.pdf"
        #
        # workbook.save(temp_excel_path)
        #
        # workbook_aspose = Workbook(temp_excel_path)
        # workbook_aspose.save(temp_pdf_path, SaveFormat.PDF)
        #
        # reader = PdfReader(temp_pdf_path)
        # writer = PdfWriter()
        #
        # pages_to_remove = [i for i in range(1, 55)]
        #
        # for i in range(len(reader.pages)):
        #     if i not in pages_to_remove:
        #         writer.add_page(reader.pages[i])
        #
        # with open(temp_modified_pdf_path, "wb") as output_pdf:
        #     writer.write(output_pdf)
        #
        # with open(temp_modified_pdf_path, "rb") as pdf_file:
        #     response = HttpResponse(pdf_file.read(), content_type="application/pdf")
        #     if watch_document:
        #         response["Content-Disposition"] = "inline; filename=invoice.pdf"
        #     else:
        #         response["Content-Disposition"] = "attachment; filename=invoice.pdf"
        #
        # os.remove(temp_excel_path)
        # os.remove(temp_pdf_path)
        # os.remove(temp_modified_pdf_path)
        #
        # return response

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=pko.xlsx"
    workbook.save(response)
    return response
