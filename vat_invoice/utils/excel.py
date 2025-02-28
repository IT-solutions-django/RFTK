import convertapi
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.worksheet.table import Table, TableStyleInfo
from aspose.cells import Workbook, HtmlSaveOptions, SaveFormat, License
from bs4 import BeautifulSoup
import aspose.cells as cells
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime
from PyPDF2 import PdfReader, PdfWriter
import os


def html_to_excel():
    html_file = "vat_invoice/utils/updated_file.html"

    html_load_options = cells.HtmlLoadOptions()

    workbook = cells.Workbook(html_file, html_load_options)

    workbook.save("vat_invoice/utils/output.xlsx")


def change_html(count_rows):
    with open("vat_invoice/utils/Счет_на_оплату.html", "r",
              encoding="utf-8") as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, "html.parser")

    element_to_remove = soup.find("div",
                                  string=lambda text: text and "Evaluation Only. Created with Aspose.Cells" in text)

    if element_to_remove:
        element_to_remove.decompose()

    target_row = soup.find("td", string=lambda
        text: text and "Кола" in text)

    if target_row:
        parent_row = target_row.find_parent("tr")

        if parent_row:
            num_rows_to_add = count_rows - 1

            if num_rows_to_add >= 1:
                for _ in range(num_rows_to_add):
                    new_row = BeautifulSoup(str(parent_row), "html.parser")

                    parent_row.insert_after(new_row)

            with open("vat_invoice/utils/updated_file.html", "w",
                      encoding="utf-8") as file:
                file.write(str(soup))


def excel_to_html():
    file_path = "vat_invoice/utils/Счет-фактура №2 от 01.02.2025 г..xlsx"
    workbook = Workbook(file_path)

    save_options = HtmlSaveOptions()
    save_options.export_active_worksheet_only = True
    save_options.export_images_as_base64 = True

    html_file = "vat_invoice/utils/Счет_на_оплату.html"
    workbook.save(html_file, save_options)


def create_vat_invoice_excel(data, formset_data, pdf=False, watch_document=False):
    excel_to_html()
    change_html(len(formset_data))
    html_to_excel()

    file_path = 'vat_invoice/utils/output.xlsx'

    workbook = openpyxl.load_workbook(file_path)
    workbook.remove(workbook["Evaluation Warning"])

    sheet = workbook["Счет-фактура"]

    for row in sheet.iter_rows():
        if row[0].row in [7, 9, 10, 11, 12]:
            sheet.row_dimensions[row[0].row].height = 22

    start_col = 208
    num_cols = 790

    sheet.delete_cols(start_col, num_cols)

    sheet['A2'] = 'Счет-фактура'
    sheet['AF2'] = data['name']

    date_str = str(data['date'])
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d-%m-%Y")
    sheet['BF2'] = formatted_date

    sheet['AF6'] = data['organization'].naming
    sheet['AF7'] = data['organization'].address
    sheet['AF8'] = f'{data["organization"].inn} / {data["organization"].kpp}'
    if data["shipper"]:
        sheet['AF9'] = f'{data["shipper"].naming}, {data["shipper"].address}'
    else:
        sheet['AF9'] = ''
    if data["consignee"]:
        sheet['AF10'] = f'{data["consignee"].naming}, {data["consignee"].address}'
    else:
        sheet['AF10'] = ''

    sheet['AF11'] = data['payment_document']
    sheet['AJ12'] = data['shipping_document']

    sheet['DG6'] = f'{data["counterparty"].naming}'
    sheet['DG7'] = f'{data["counterparty"].address}'
    sheet['DG8'] = f'{data["counterparty"].inn} / {data["counterparty"].kpp}'
    sheet['DG9'] = f'{data["currency"]}'
    sheet['DV10'] = data['state_ID_contract']

    start_table_row = 25
    total_sum = 0
    total_sum_nds = 0
    total_sum_with_nds = 0

    if int(data['nds']) > 0 and int(data['nds']):
        nds = int(data['nds'])
    else:
        nds = 0

    for idx, table_data in enumerate(formset_data, 1):
        sheet.row_dimensions[start_table_row + idx].height = 22

        total_sum += table_data["amount"]

        sheet[f'A{start_table_row + idx}'] = f'{idx}'
        sheet[f'E{start_table_row + idx}'] = f'{table_data["name"]}'
        if table_data["product_type_code"]:
            sheet[f'AD{start_table_row + idx}'] = f'{table_data["product_type_code"]}'
        else:
            sheet[f'AD{start_table_row + idx}'] = ''
        sheet[f'AK{start_table_row + idx}'] = ''
        sheet[f'AP{start_table_row + idx}'] = f'{table_data["unit_of_measurement"]}'
        sheet[f'BD{start_table_row + idx}'] = f'{table_data["quantity"]}'
        sheet[f'BL{start_table_row + idx}'] = f'{table_data["price"]}'
        sheet[f'BW{start_table_row + idx}'] = f'{round(float(table_data["quantity"]) * float(table_data["price"]), 2)}'
        total_sum_with_nds += float(table_data["quantity"]) * float(table_data["price"])

        if table_data["excise"]:
            sheet[f'CL{start_table_row + idx}'] = f'{table_data["excise"]}'
        else:
            sheet[f'CL{start_table_row + idx}'] = 'Без акциза'
        if data["nds"] == '-1' or data["nds"] == -1:
            sheet[f'CS{start_table_row + idx}'] = 'Без НДС'
        else:
            sheet[f'CS{start_table_row + idx}'] = f'{data["nds"]}%'
        if nds > 0:
            sheet[
                f'CY{start_table_row + idx}'] = f'{round(float(table_data["quantity"]) * float(table_data["price"]) * nds * 0.01, 2)}'
            total_sum_nds += float(table_data["quantity"]) * float(table_data["price"]) * nds * 0.01
        else:
            sheet[f'CY{start_table_row + idx}'] = f'{0}'

        sheet[f'DN{start_table_row + idx}'] = f'{table_data["amount"]}'

        if table_data["country"]:
            sheet[f'EC{start_table_row + idx}'] = ''
            sheet[f'EK{start_table_row + idx}'] = f'{table_data["country"]}'
        else:
            sheet[f'EC{start_table_row + idx}'] = ''
            sheet[f'EK{start_table_row + idx}'] = ''

        if table_data["number_GTD"]:
            sheet[f'ES{start_table_row + idx}'] = f'{table_data["number_GTD"]}'
        else:
            sheet[f'ES{start_table_row + idx}'] = ''

    sheet[f'BW{start_table_row + len(formset_data) + 1}'] = f'{round(total_sum_with_nds, 2)}'
    sheet[f'CY{start_table_row + len(formset_data) + 1}'] = f'{round(total_sum_nds, 2)}'
    sheet[f'DN{start_table_row + len(formset_data) + 1}'] = f'{float(total_sum)}'

    sheet[f'AW{start_table_row + len(formset_data) + 3}'] = data['organization'].supervisor
    sheet[f'EB{start_table_row + len(formset_data) + 3}'] = data['organization'].accountant

    if data['organization'].stamp and data['is_stamp']:
        image_file = data['organization'].stamp
        img = Image(BytesIO(image_file.read()))

        img.width = 45 * 2.83
        img.height = 45 * 2.83

        sheet.add_image(img, f"CY{start_table_row + len(formset_data) + 6}")

    if data['organization'].signature and data['is_stamp']:
        image_file = data['organization'].signature

        image_data = image_file.read()

        img_stream = BytesIO(image_data)

        img1 = Image(img_stream)
        img1.width = 70
        img1.height = 30
        sheet.add_image(img1, f"AH{start_table_row + len(formset_data) + 3}")

    if pdf:
        convertapi.api_credentials = 'secret_VEJPjELYZzhUihM6'

        temp_excel_path = "vat_invoice/utils/invoice.xlsx"
        temp_modified_pdf_path = "vat_invoice/utils/invoice_modified.pdf"

        workbook.save(temp_excel_path)

        temp_pdf_path = convertapi.convert('pdf', {
            'File': temp_excel_path,
        }, from_format='xls').save_files('vat_invoice/utils')[0]

        reader = PdfReader(temp_pdf_path)
        writer = PdfWriter()

        pages_to_remove = [i for i in range(1, 55)]

        for i in range(len(reader.pages)):
            writer.add_page(reader.pages[i])
            # if i not in pages_to_remove:
            #     writer.add_page(reader.pages[i])

        with open(temp_modified_pdf_path, "wb") as output_pdf:
            writer.write(output_pdf)

        with open(temp_modified_pdf_path, "rb") as pdf_file:
            response = HttpResponse(pdf_file.read(), content_type="application/pdf")
            if watch_document:
                response["Content-Disposition"] = "inline; filename=Счет-фактура.pdf"
            else:
                response["Content-Disposition"] = "attachment; filename=Счет-фактура.pdf"

        os.remove(temp_excel_path)
        os.remove(temp_pdf_path)
        os.remove(temp_modified_pdf_path)

        return response

        # temp_excel_path = "vat_invoice/utils/invoice.xlsx"
        # temp_pdf_path = "vat_invoice/utils/invoice.pdf"
        # temp_modified_pdf_path = "vat_invoice/utils/invoice_modified.pdf"
        #
        # workbook.save(temp_excel_path)
        #
        # workbook_aspose = Workbook(temp_excel_path)
        # workbook_aspose.save(temp_pdf_path, SaveFormat.PDF)
        #
        # reader = PdfReader(temp_pdf_path)
        # writer = PdfWriter()
        #
        # pages_to_remove = [i for i in range(1, 55)]
        #
        # for i in range(len(reader.pages)):
        #     if i not in pages_to_remove:
        #         writer.add_page(reader.pages[i])
        #
        # with open(temp_modified_pdf_path, "wb") as output_pdf:
        #     writer.write(output_pdf)
        #
        # with open(temp_modified_pdf_path, "rb") as pdf_file:
        #     response = HttpResponse(pdf_file.read(), content_type="application/pdf")
        #     if watch_document:
        #         response["Content-Disposition"] = "inline; filename=invoice.pdf"
        #     else:
        #         response["Content-Disposition"] = "attachment; filename=invoice.pdf"
        #
        # os.remove(temp_excel_path)
        # os.remove(temp_pdf_path)
        # os.remove(temp_modified_pdf_path)
        #
        # return response

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=Счет-фактура.xlsx"
    workbook.save(response)
    return response
